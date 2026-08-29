from datetime import datetime

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import preparar_archivo_datos
import publicaciones
from trazabilidad import necesita_reprocesamiento
from publicaciones import (
    COLUMNAS_PUBLICACIONES,
    crear_registro_publicacion,
    debe_procesar_publicacion,
    normalizar_publicaciones,
    puede_reutilizar_publicacion,
    publicaciones_desde_oposiciones,
    registrar_publicacion,
)


def _enlace(publicacion_id="BOE-A-2026-10463"):
    return f"https://www.boe.es/diario_boe/txt.php?id={publicacion_id}"


def test_registro_persiste_sumario_y_trazabilidad_sin_tocar_convocatorias():
    registro = crear_registro_publicacion(
        _enlace(), "22 de agosto de 2026", "Título HTML", 2,
        titulo_sumario="Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria.",
        departamento_boe="Administración Local",
    )
    assert registro["Titulo_original"].startswith("Resolución del Ayuntamiento")
    assert registro["Departamento_BOE"] == "Administración Local"
    assert (registro["Administracion_resuelta"], registro["Familia_administrativa"],
            registro["Estado_resolucion"], registro["Confianza_resolucion"],
            registro["Version_resolucion"]) == ("Ayuntamiento de Ciudad Real", "AYUNTAMIENTO", "RESUELTA", "ALTA", "1")


def test_resolucion_ambigua_no_inventa_administracion():
    registro = crear_registro_publicacion(
        _enlace(), "22 de agosto de 2026",
        "Resolución del Ayuntamiento de A, referente a la convocatoria del Consorcio B.", 0,
    )
    assert registro["Administracion_resuelta"] == ""
    assert (registro["Estado_resolucion"], registro["Confianza_resolucion"]) == ("AMBIGUA", "AMBIGUA")


def _publicaciones_con_version(version):
    return pd.DataFrame(
        [
            {
                "Publicacion_ID": "BOE-A-2026-10463",
                "Version_extractor": version,
            }
        ]
    )


def test_publicacion_no_presente_en_busquedas_se_procesa_normalmente():
    assert debe_procesar_publicacion(
        _enlace(), set(), _enlace(), _publicaciones_con_version("1")
    )


@pytest.mark.parametrize(("version", "esperado"), [("1", False), ("legacy", True)])
def test_publicacion_procesada_depende_de_su_version(version, esperado):
    codigo = _enlace()

    assert (
        debe_procesar_publicacion(
            codigo, {codigo}, _enlace(), _publicaciones_con_version(version)
        )
        is esperado
    )


def test_publicacion_con_version_anterior_se_reprocesa(monkeypatch):
    codigo = _enlace()
    monkeypatch.setattr(
        publicaciones,
        "necesita_reprocesamiento",
        lambda version: necesita_reprocesamiento(version, "2"),
    )

    assert debe_procesar_publicacion(
        codigo, {codigo}, _enlace(), _publicaciones_con_version("1")
    )


def test_publicacion_sin_id_valido_mantiene_exclusion_de_busquedas():
    enlace = "https://www.boe.es/diario_boe/txt.php?id=no-valido"

    assert not debe_procesar_publicacion(enlace, {enlace}, enlace, pd.DataFrame())


def _registro_reutilizable(version="1", estado="con_coincidencias", numero=1):
    return pd.DataFrame(
        [
            {
                "Publicacion_ID": "BOE-A-2026-10463",
                "Version_extractor": version,
                "Estado_analisis": estado,
                "Coincidencias": numero,
            }
        ]
    )


@pytest.mark.parametrize("version", [None, "", "legacy", "invalida"])
def test_publicacion_con_version_desconocida_no_se_reutiliza(version):
    assert not puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable(version),
        pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-10463"]}),
    )


def test_publicacion_desconocida_o_sin_id_no_se_reutiliza():
    oposiciones = pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-10463"]})

    assert not puede_reutilizar_publicacion(
        "BOE-A-2026-10463", pd.DataFrame(), oposiciones
    )
    assert not puede_reutilizar_publicacion(
        None, _registro_reutilizable(), oposiciones
    )


def test_publicacion_con_version_anterior_no_se_reutiliza():
    assert not puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable("1"),
        pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-10463"]}),
        version_actual="2",
    )


@pytest.mark.parametrize("version", ["1", "2"])
def test_publicacion_con_coincidencias_y_datos_locales_se_reutiliza(version):
    assert puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable(version),
        pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-10463"]}),
    )


def test_version_posterior_se_protege_aunque_no_haya_filas_locales():
    assert puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable("2"),
        pd.DataFrame(),
    )


def test_publicacion_sin_coincidencias_se_reutiliza_sin_filas_locales():
    assert puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable("1", "sin_coincidencias", 0),
        pd.DataFrame(),
    )


def test_publicacion_inconsistente_no_se_reutiliza():
    assert not puede_reutilizar_publicacion(
        "BOE-A-2026-10463",
        _registro_reutilizable("1", "con_coincidencias", 2),
        pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-otra"]}),
    )


def test_comprobar_reutilizacion_no_modifica_los_dataframes():
    publicaciones_df = _registro_reutilizable()
    oposiciones = pd.DataFrame({"Publicacion_ID": ["BOE-A-2026-10463"]})
    copia_publicaciones = publicaciones_df.copy(deep=True)
    copia_oposiciones = oposiciones.copy(deep=True)

    puede_reutilizar_publicacion(
        "BOE-A-2026-10463", publicaciones_df, oposiciones
    )

    pd.testing.assert_frame_equal(publicaciones_df, copia_publicaciones)
    pd.testing.assert_frame_equal(oposiciones, copia_oposiciones)


def test_crea_registros_con_y_sin_coincidencias():
    momento = datetime(2026, 8, 22, 12, 34, 56)

    con = crear_registro_publicacion(
        _enlace(), "BOE núm. 10, de 22 de agosto de 2026", "Título completo", 3, momento
    )
    sin = crear_registro_publicacion(
        _enlace("BOE-A-2026-10464"),
        "22 de agosto de 2026",
        "Otro título",
        0,
        momento,
    )

    assert list(con) == COLUMNAS_PUBLICACIONES
    assert con["Fecha_BOE"] == "22 de agosto de 2026"
    assert con["Fecha_ultimo_analisis"] == "2026-08-22 12:34:56"
    assert con["Estado_analisis"] == "con_coincidencias"
    assert con["Coincidencias"] == 3
    assert sin["Estado_analisis"] == "sin_coincidencias"
    assert sin["Coincidencias"] == 0


def test_actualiza_publicacion_sin_duplicar_y_preserva_datos_validos():
    existente = pd.DataFrame(
        [
            {
                "Publicacion_ID": "BOE-A-2026-10463",
                "Enlace": _enlace(),
                "Fecha_BOE": "21 de agosto de 2026",
                "Titulo_original": "Título conservado",
                "Fecha_ultimo_analisis": "2026-08-21 10:00:00",
                "Version_extractor": "legacy",
                "Estado_analisis": "con_coincidencias",
                "Coincidencias": 1,
            }
        ]
    )
    nuevo = {
        "Publicacion_ID": "BOE-A-2026-10463",
        "Enlace": "",
        "Fecha_BOE": pd.NA,
        "Titulo_original": "",
        "Fecha_ultimo_analisis": "2026-08-22 12:00:00",
        "Version_extractor": "1",
        "Estado_analisis": "sin_coincidencias",
        "Coincidencias": 0,
    }
    copia = existente.copy(deep=True)

    resultado = registrar_publicacion(existente, nuevo)

    assert len(resultado) == 1
    fila = resultado.iloc[0]
    assert fila["Enlace"] == _enlace()
    assert fila["Fecha_BOE"] == "21 de agosto de 2026"
    assert fila["Titulo_original"] == "Título conservado"
    assert fila["Fecha_ultimo_analisis"] == "2026-08-22 12:00:00"
    assert fila["Version_extractor"] == "1"
    assert fila["Estado_analisis"] == "sin_coincidencias"
    assert fila["Coincidencias"] == 0
    pd.testing.assert_frame_equal(existente, copia)


def test_consolida_duplicados_preexistentes_sin_perder_datos_validos():
    duplicadas = pd.DataFrame(
        [
            {
                "Publicacion_ID": "BOE-A-2026-10463",
                "Enlace": _enlace(),
                "Fecha_BOE": "22 de agosto de 2026",
                "Titulo_original": "Título conservado",
                "Fecha_ultimo_analisis": "2026-08-21 10:00:00",
                "Version_extractor": "legacy",
                "Estado_analisis": "con_coincidencias",
                "Coincidencias": 1,
            },
            {
                "Publicacion_ID": "BOE-A-2026-10463",
                "Enlace": "",
                "Fecha_BOE": "",
                "Titulo_original": "",
                "Fecha_ultimo_analisis": "2026-08-22 10:00:00",
                "Version_extractor": "1",
                "Estado_analisis": "sin_coincidencias",
                "Coincidencias": 0,
            },
        ]
    )

    resultado = normalizar_publicaciones(duplicadas)

    assert len(resultado) == 1
    assert resultado.loc[0, "Titulo_original"] == "Título conservado"
    assert resultado.loc[0, "Fecha_ultimo_analisis"] == "2026-08-22 10:00:00"


def test_reconstruye_historico_una_fila_por_publicacion_y_cuenta_coincidencias():
    oposiciones = pd.DataFrame(
        {
            "Puesto": ["Ingeniero", "Arquitecto"],
            "Enlace": [_enlace(), _enlace()],
            "Fecha_boe": ["22 de agosto de 2026"] * 2,
        }
    )

    resultado = publicaciones_desde_oposiciones(oposiciones)

    assert len(resultado) == 1
    fila = resultado.iloc[0]
    assert fila["Publicacion_ID"] == "BOE-A-2026-10463"
    assert fila["Coincidencias"] == 2
    assert fila["Estado_analisis"] == "con_coincidencias"
    assert fila["Version_extractor"] == "legacy"
    assert fila["Titulo_original"] == ""
    assert pd.isna(fila["Fecha_ultimo_analisis"])


def test_carga_crea_publicaciones_solo_desde_oposiciones(monkeypatch, tmp_path):
    ruta = tmp_path / "BOE-oposiciones.xlsx"
    busquedas = pd.DataFrame(
        {"Código": [_enlace(), _enlace("BOE-A-2026-99999")]}
    )
    oposiciones = pd.DataFrame(
        {"Puesto": ["Ingeniero"], "Enlace": [_enlace()], "Fecha_boe": ["22 de agosto de 2026"]}
    )
    errores = pd.DataFrame(columns=["Fecha", "Tipo de error", "Enlace Web"])
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        busquedas.to_excel(writer, sheet_name="Búsquedas", index=False)
        oposiciones.to_excel(writer, sheet_name="Oposiciones", index=False)
        errores.to_excel(writer, sheet_name="Log-errores", index=False)
    monkeypatch.chdir(tmp_path)

    hojas = preparar_archivo_datos.preparar_excel_y_dataframes()

    assert hojas["Publicaciones"]["Publicacion_ID"].tolist() == ["BOE-A-2026-10463"]


def test_guardado_atomico_incluye_publicaciones_y_conserva_hojas(monkeypatch, tmp_path):
    ruta = tmp_path / "BOE-oposiciones.xlsx"
    libro = Workbook()
    libro.active.title = "Búsquedas"
    libro.create_sheet("Oposiciones")
    libro.create_sheet("Log-errores")
    libro.save(ruta)
    monkeypatch.chdir(tmp_path)
    publicaciones = pd.DataFrame(
        [crear_registro_publicacion(_enlace(), "22 de agosto de 2026", "Título", 1)]
    )

    preparar_archivo_datos.guardar_excel(
        pd.DataFrame({"Puesto": ["Ingeniero"]}),
        pd.DataFrame({"Código": [_enlace()]}),
        pd.DataFrame(columns=["Fecha", "Tipo de error", "Enlace Web"]),
        publicaciones,
    )

    comprobacion = load_workbook(ruta, read_only=True)
    assert comprobacion.sheetnames == [
        "Búsquedas",
        "Oposiciones",
        "Log-errores",
        "Publicaciones",
        "Cobertura",
    ]
    assert comprobacion["Búsquedas"].sheet_state == "hidden"
    comprobacion.close()
    guardadas = pd.read_excel(ruta, sheet_name="Publicaciones")
    assert guardadas["Publicacion_ID"].tolist() == ["BOE-A-2026-10463"]
