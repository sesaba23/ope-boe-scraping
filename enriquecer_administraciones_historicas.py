"""Enriquecimiento experimental de Administraciones históricas desde sumarios BOE.

Solo construye un catálogo JSON y calcula una propuesta en memoria. No escribe
el libro Excel, no descarga XML y no modifica los estados del cargador.
"""
import argparse
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import sys
import tempfile
import time

import pandas as pd
from openpyxl import load_workbook

from tqdm import tqdm

from boe_api import extraer_publicaciones_2b_api, obtener_sumario_api
from diagnostico_administraciones_historicas import cargar_historicos
from diagnostico_administraciones_historicas import cargar_catalogo, crear_indice_municipios, resolver_entidad
from resolucion_administraciones import (
    _variantes_provincia,
    cargar_alias_municipios, cargar_capitales_provinciales,
    cargar_sedes_administraciones, es_generica, extraer_administraciones_titulo,
    resolver_sedes,
)
from mapa_plazas import normalizar_nombre_municipal
from fechas import convertir_fecha
from preparar_archivo_datos import bloqueo_excel


VERSION_CATALOGO = 1
VERSION_HISTORICA = "historico-experimental-2004"
RUTA_CATALOGO = "datos/catalogo_administraciones_historicas.json"
INTERVALO_GUARDADO_FECHAS = 25
TRABAJADORES_SUMARIO = 8


def firma_excel(ruta):
    ruta = Path(ruta); estado = ruta.stat()
    return {"sha256": hashlib.sha256(ruta.read_bytes()).hexdigest(), "tamano": estado.st_size,
            "mtime_ns": estado.st_mtime_ns}


def guardar_json_atomico(ruta, datos):
    ruta = Path(ruta); ruta.parent.mkdir(parents=True, exist_ok=True)
    contenido = json.dumps(datos, ensure_ascii=False, indent=2) + "\n"
    json.loads(contenido)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=ruta.parent, delete=False) as temporal:
        temporal.write(contenido); temporal.flush(); os.fsync(temporal.fileno()); nombre = temporal.name
    os.replace(nombre, ruta)


def cargar_o_crear_catalogo(ruta, fechas):
    ruta = Path(ruta)
    if ruta.exists():
        catalogo = json.loads(ruta.read_text(encoding="utf-8"))
        if catalogo.get("version_formato") != VERSION_CATALOGO:
            raise ValueError("Formato de catálogo histórico incompatible")
        # La primera ejecución experimental podía haber recibido Fecha_BOE en
        # castellano y truncarla. Conservamos exclusivamente las respuestas de
        # fechas ISO válidas y descartamos esas claves técnicas no consultables.
        validas = set(fechas)
        catalogo["fechas_objetivo"] = fechas
        catalogo["fechas_completadas"] = [x for x in catalogo.get("fechas_completadas", []) if x in validas]
        catalogo["errores_por_fecha"] = {x: e for x, e in catalogo.get("errores_por_fecha", {}).items() if x in validas}
        return catalogo
    return {"version_formato": VERSION_CATALOGO, "fecha_creacion": datetime.now().isoformat(timespec="seconds"),
            "fecha_actualizacion": datetime.now().isoformat(timespec="seconds"),
            "fechas_objetivo": fechas, "fechas_completadas": [], "errores_por_fecha": {},
            "publicaciones": {}}


def fechas_historicas(publicaciones):
    def normalizar(valor):
        texto = str(valor).strip()
        try:
            return datetime.fromisoformat(texto[:10]).date().isoformat()
        except ValueError:
            return convertir_fecha(texto.casefold()).date().isoformat()
    return sorted({normalizar(x) for x in publicaciones["Fecha_BOE"].dropna()})


class ProgresoFechas:
    def __init__(self, total, acumuladas=0, publicaciones=0, errores=0, stream=None):
        self.total, self.stream, self.actual, self.inicio = total, stream or sys.stdout, 0, time.monotonic()
        self.tty = bool(getattr(self.stream, "isatty", lambda: False)())
        self.intervalo = max(1, total // 10) if total else 1
        self.acumuladas, self.publicaciones, self.errores = acumuladas, publicaciones, errores
        self.barra = tqdm(total=total, desc="Recuperando sumarios BOE", dynamic_ncols=True,
                          file=self.stream, disable=not self.tty) if self.tty else None
        if not self.tty:
            print(f"Sumarios pendientes: 0/{total}", file=self.stream)

    def actualizar(self):
        self.actual += 1
        transcurrido = time.monotonic() - self.inicio
        velocidad = self.actual / transcurrido if transcurrido else 0
        eta = "calculando..." if self.actual < 2 or not velocidad else f"{(self.total-self.actual)/velocidad:.0f}s"
        if self.barra:
            self.barra.update(1); self.barra.set_postfix_str(f"acumuladas={self.acumuladas + self.actual}; publicaciones={self.publicaciones}; errores={self.errores}")
        elif self.actual == self.total or self.actual % self.intervalo == 0:
            print(f"Progreso: {self.actual}/{self.total}; {velocidad:.1f} fechas/s; ETA {eta}", file=self.stream)

    def cerrar(self):
        if self.barra:
            self.barra.close()


class ProgresoAnalisis:
    def __init__(self, total, stream=None):
        self.total, self.stream = total, stream or sys.stdout
        self.tty = bool(getattr(self.stream, "isatty", lambda: False)())
        self.actual = self.resueltas = self.ambiguas = self.no_resueltas = 0
        self.intervalo = max(1, total // 10) if total else 1
        self.barra = tqdm(total=total, desc="Analizando administraciones", dynamic_ncols=True, file=self.stream, disable=not self.tty) if self.tty else None

    def actualizar(self, confianza=None):
        self.actual += 1
        if confianza == "ALTA": self.resueltas += 1
        elif confianza == "AMBIGUA": self.ambiguas += 1
        elif confianza == "NO_RESUELTA": self.no_resueltas += 1
        if self.barra:
            self.barra.update(1); self.barra.set_postfix_str(f"resueltas={self.resueltas}; ambiguas={self.ambiguas}; no_resueltas={self.no_resueltas}")
        elif self.actual == self.total or self.actual % self.intervalo == 0:
            print(f"Análisis: {self.actual}/{self.total}; resueltas={self.resueltas}; ambiguas={self.ambiguas}; no_resueltas={self.no_resueltas}", file=self.stream)

    def cerrar(self):
        if self.barra: self.barra.close()


def construir_catalogo(publicaciones=None, ruta=RUTA_CATALOGO, obtener=obtener_sumario_api, stream=None):
    """Completa solo fechas pendientes y persiste tras cada respuesta válida."""
    ruta = Path(ruta)
    existente = json.loads(ruta.read_text(encoding="utf-8")) if ruta.exists() else None
    if existente and existente.get("publicacion_ids_objetivo"):
        fechas = existente["fechas_objetivo"]
        ids_objetivo = set(existente["publicacion_ids_objetivo"])
    else:
        if publicaciones is None:
            raise ValueError("El primer catálogo necesita Publicaciones históricas")
        fechas = fechas_historicas(publicaciones)
        ids_objetivo = set(publicaciones["Publicacion_ID"].dropna().astype(str))
    catalogo = cargar_o_crear_catalogo(ruta, fechas)
    catalogo.setdefault("publicacion_ids_objetivo", sorted(ids_objetivo))
    # Persiste inmediatamente una migración de claves antiguas, antes de la
    # siguiente petición, para que la reanudación no vuelva a verlas.
    guardar_json_atomico(ruta, catalogo)
    pendientes = [x for x in fechas if x not in set(catalogo.get("fechas_completadas", []))]
    salida = stream or sys.stdout
    print(f"Catálogo BOE:\nFechas totales: {len(fechas):,}\nYa completadas: {len(catalogo.get('fechas_completadas', [])):,}\nPendientes: {len(pendientes):,}", file=salida)
    progreso = ProgresoFechas(len(pendientes), len(catalogo.get("fechas_completadas", [])), len(catalogo.get("publicaciones", {})), len(catalogo.get("errores_por_fecha", {})), salida)
    modificadas = 0
    def consultar(fecha):
        return fecha, extraer_publicaciones_2b_api(obtener(fecha))
    with ThreadPoolExecutor(max_workers=TRABAJADORES_SUMARIO) as ejecutor:
      futuros = {ejecutor.submit(consultar, fecha): fecha for fecha in pendientes}
      for futuro in as_completed(futuros):
        fecha = futuros[futuro]
        try:
            _, resultado = futuro.result()
            for item in resultado.get("publicaciones", []):
                publicacion_id = item.get("Publicacion_ID")
                if publicacion_id in ids_objetivo:
                    catalogo["publicaciones"][publicacion_id] = {
                        "titulo": item.get("titulo", ""), "departamento": item.get("departamento", ""),
                        "url_html": item.get("url_html", ""), "url_xml": item.get("url_xml", ""),
                        "fecha_sumario": fecha,
                    }
            progreso.publicaciones = len(catalogo["publicaciones"])
            catalogo["fechas_completadas"].append(fecha)
            catalogo["errores_por_fecha"].pop(fecha, None)
        except KeyboardInterrupt:
            catalogo["fecha_actualizacion"] = datetime.now().isoformat(timespec="seconds")
            guardar_json_atomico(ruta, catalogo); progreso.cerrar(); raise
        except Exception as error:
            catalogo["errores_por_fecha"][fecha] = str(error)
            progreso.errores = len(catalogo["errores_por_fecha"])
        catalogo["fecha_actualizacion"] = datetime.now().isoformat(timespec="seconds")
        modificadas += 1
        if modificadas % INTERVALO_GUARDADO_FECHAS == 0:
            guardar_json_atomico(ruta, catalogo)
        progreso.actualizar()
    progreso.cerrar()
    guardar_json_atomico(ruta, catalogo)
    return catalogo


def propuesta_dry_run(publicaciones, oposiciones, catalogo, stream=None):
    filas_por_id = Counter(oposiciones["Publicacion_ID"].dropna().astype(str))
    candidatas = oposiciones[oposiciones["Administración"].map(es_generica)]
    propuestas, familias, detecciones = [], Counter(), {}
    progreso = ProgresoAnalisis(len(oposiciones), stream)
    for fila in oposiciones.to_dict(orient="records"):
        publicacion_id = str(fila.get("Publicacion_ID") or "")
        deteccion = None
        if es_generica(fila.get("Administración")) and publicacion_id:
            if publicacion_id not in detecciones:
                detecciones[publicacion_id] = extraer_administraciones_titulo(catalogo.get("publicaciones", {}).get(publicacion_id, {}).get("titulo", ""))
            deteccion = detecciones[publicacion_id]
        progreso.actualizar(deteccion.get("confianza") if deteccion else None)
    progreso.cerrar()
    for publicacion_id in sorted(set(candidatas["Publicacion_ID"].dropna().astype(str))):
        item = catalogo.get("publicaciones", {}).get(publicacion_id, {})
        deteccion = detecciones.get(publicacion_id) or extraer_administraciones_titulo(item.get("titulo", ""))
        propuesta = {"Publicacion_ID": publicacion_id, "titulo": item.get("titulo", ""),
                     "departamento_BOE": item.get("departamento", ""), **deteccion,
                     "filas_oposiciones": int(filas_por_id[publicacion_id])}
        propuestas.append(propuesta)
        if deteccion["familia"]:
            familias[deteccion["familia"]] += propuesta["filas_oposiciones"]
    corregibles = sum(x["filas_oposiciones"] for x in propuestas if x["confianza"] == "ALTA")
    total = len(oposiciones); antes = total - len(candidatas)
    sedes = resolver_sedes(propuestas, stream)
    por_id = {x["Publicacion_ID"]: x for x in sedes["propuestas"]}
    prevision = Counter()
    for fila in candidatas.to_dict(orient="records"):
        propuesta = por_id.get(str(fila.get("Publicacion_ID") or ""), {})
        if propuesta.get("confianza") != "ALTA":
            continue
        prevision["administraciones_a_actualizar"] += 1
        if propuesta.get("confianza_sede") == "ALTA":
            if fila.get("Municipio") != propuesta.get("Municipio"):
                prevision["municipios_a_actualizar"] += 1
            if fila.get("Provincia") != propuesta.get("Provincia"):
                prevision["provincias_a_actualizar"] += 1
    return {"resumen": {
                "fechas_objetivo": len(catalogo.get("fechas_objetivo", [])),
                "fechas_consultadas": len(catalogo.get("fechas_completadas", [])),
                "errores_fecha": len(catalogo.get("errores_por_fecha", {})),
                "publicaciones_boe_recuperadas": len(catalogo.get("publicaciones", {})),
                "publicacion_id_encontrados_excel": sum(x["Publicacion_ID"] in catalogo.get("publicaciones", {}) for x in propuestas),
                "filas_oposiciones_candidatas": len(candidatas), "administraciones_resueltas": sum(x["confianza"] == "ALTA" for x in propuestas),
                "filas_potencialmente_corregibles": corregibles,
                "no_resueltas": sum(x["confianza"] == "NO_RESUELTA" for x in propuestas),
                "ambiguas": sum(x["confianza"] == "AMBIGUA" for x in propuestas),
                "cobertura_antes_pct": round(100 * antes / total, 2) if total else 0,
                "cobertura_prevista_despues_pct": round(100 * (antes + corregibles) / total, 2) if total else 0,
                "cobertura_administracion_pct": round(100 * (antes + corregibles) / total, 2) if total else 0,
                "cobertura_municipio_prevista_pct": round(100 * sedes["resumen"]["filas_completamente_geolocalizables"] / total, 2) if total else 0,
                "cobertura_provincia_prevista_pct": round(100 * sedes["resumen"]["filas_completamente_geolocalizables"] / total, 2) if total else 0,
                **sedes["resumen"],
            }, "prevision_actualizacion": dict(prevision), "familias_por_filas": dict(familias.most_common()), "propuestas": propuestas,
            "propuestas_sede": sedes["propuestas"],
            "sedes": sedes["sedes"], "no_resueltas_sedes": sedes["no_resueltas"],
            "administraciones_mas_frecuentes": Counter(x["administracion_detectada"] for x in propuestas if x["confianza"] == "ALTA").most_common(30)}


def analizar_alias_municipios(propuestas, total_base=1467):
    """Audita alias aplicados y Ayuntamientos que siguen sin municipio."""
    catalogo = cargar_catalogo()
    indice = crear_indice_municipios(catalogo)
    aliases = cargar_alias_municipios()
    por_codigo = {str(x["Codigo_INE"]).zfill(5): x for valores in aliases.values() for x in valores}
    aplicados = Counter()
    pendientes = Counter()
    detalles = {}
    for propuesta in propuestas:
        if propuesta.get("familia") != "AYUNTAMIENTO":
            continue
        metodo = propuesta.get("metodo_resolucion", "")
        if "CATALOGO_ALIAS_MUNICIPIOS_" in metodo:
            codigo = metodo.rsplit("_", 1)[-1].zfill(5)
            aplicados[codigo] += propuesta["filas_oposiciones"]
            continue
        if propuesta.get("confianza_sede") == "ALTA":
            continue
        administracion = propuesta.get("administracion_detectada", "")
        entidad = re.sub(r"^Ayuntamiento de\s+", "", administracion, flags=re.I).split(",", 1)[0].strip()
        municipio, provincia = _entidad_y_provincia(entidad)
        candidatos = indice.get(normalizar_nombre_municipal(municipio), [])
        if provincia:
            variantes = _variantes_provincia(provincia)
            candidatos = [x for x in candidatos if _variantes_provincia(x["Provincia"]) & variantes]
        if "AMBIGUO" in metodo:
            causa = "AMBIGUEDAD_REAL"
        elif any(x in administracion.casefold() for x in (
                "referente a la convocatoria", "organismo autónomo", "patronato", "fundación", "residencia")):
            causa = "TEXTO_ADICIONAL_O_EXTRACCION_INCORRECTA"
        else:
            causa = "MUNICIPIO_NO_IDENTIFICABLE"
        clave = (administracion, municipio, provincia or "", propuesta.get("familia", ""))
        pendientes[clave] += propuesta["filas_oposiciones"]
        detalles[clave] = {
            "Administracion": administracion, "municipio_extraido": municipio,
            "provincia_extraida": provincia or "", "familia": propuesta.get("familia", ""),
            "candidatos_exactos": sorted({f"{x['Población']} ({x['Provincia']})" for x in candidatos}),
            "causa_probable": causa,
        }
    ranking_alias = []
    acumulado = 0
    for posicion, (codigo, filas) in enumerate(aplicados.most_common(), start=1):
        acumulado += filas
        alias = por_codigo[codigo]
        ranking_alias.append({"posicion": posicion, "Alias": alias["Alias"], "Provincia": alias["Provincia"],
                              "Municipio_oficial": alias["Municipio_oficial"], "Codigo_INE": codigo,
                              "filas_oposiciones": filas, "filas_acumuladas": acumulado})
    umbrales = {}
    for porcentaje in (50, 75, 90, 95):
        objetivo = total_base * porcentaje / 100
        fila = next((x for x in ranking_alias if x["filas_acumuladas"] >= objetivo), None)
        umbrales[str(porcentaje)] = ({"aliases_necesarios": fila["posicion"], "filas_cubiertas": fila["filas_acumuladas"]}
                                     if fila else {"aliases_necesarios": None, "filas_cubiertas": acumulado})
    ranking_pendientes = []
    for posicion, (clave, filas) in enumerate(pendientes.most_common(), start=1):
        ranking_pendientes.append({"posicion": posicion, **detalles[clave], "filas_oposiciones": filas})
    return {"alias_incorporados": len(por_codigo), "filas_resueltas_por_alias": sum(aplicados.values()),
            "ranking_alias": ranking_alias, "umbrales_cobertura": umbrales,
            "ranking_municipios_no_resueltos": ranking_pendientes,
            "causas_restantes": dict(Counter(x["causa_probable"] for x in ranking_pendientes for _ in range(x["filas_oposiciones"]))) }


def escribir_diagnostico_alias(resultado, directorio="informes/diagnostico_administraciones_historicas"):
    destino = Path(directorio); destino.mkdir(parents=True, exist_ok=True)
    (destino / "diagnostico_alias_municipios.json").write_text(
        json.dumps(resultado, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lineas = ["# Diagnóstico de alias municipales", "", "## Resumen", ""]
    lineas += [f"- {clave}: {valor}" for clave, valor in resultado.items()
               if clave not in {"ranking_alias", "ranking_municipios_no_resueltos"}]
    lineas += ["", "## Umbrales de cobertura sobre 1.467 filas", "",
               "| Objetivo | Alias necesarios | Filas cubiertas |", "|---:|---:|---:|"]
    lineas += [f"| {objetivo}% | {datos['aliases_necesarios'] if datos['aliases_necesarios'] is not None else 'No alcanzable'} | {datos['filas_cubiertas']} |"
               for objetivo, datos in resultado["umbrales_cobertura"].items()]
    lineas += ["", "## Ranking completo de alias aplicados", "",
               "| # | Alias | Provincia | Municipio oficial | Código INE | Filas |", "|---:|---|---|---|---|---:|"]
    lineas += [f"| {x['posicion']} | {x['Alias']} | {x['Provincia']} | {x['Municipio_oficial']} | {x['Codigo_INE']} | {x['filas_oposiciones']} |"
               for x in resultado["ranking_alias"]]
    lineas += ["", "## Ranking completo pendiente", "",
               "| # | Administración | Municipio extraído | Provincia | Filas | Candidatos exactos | Causa |",
               "|---:|---|---|---|---:|---|---|"]
    lineas += [f"| {x['posicion']} | {x['Administracion']} | {x['municipio_extraido']} | {x['provincia_extraida']} | {x['filas_oposiciones']} | {'; '.join(x['candidatos_exactos'])} | {x['causa_probable']} |"
               for x in resultado["ranking_municipios_no_resueltos"]]
    (destino / "diagnostico_alias_municipios.md").write_text("\n".join(lineas) + "\n", encoding="utf-8")


def analizar_pendientes_sede(propuestas):
    """Ordena las entidades que requieren una sede explícita sin inferirla."""
    pendientes = Counter()
    causas = {}
    familias_supramunicipales = {
        "MINISTERIO", "CONSEJERIA", "GOBIERNO", "JUNTA", "GENERALITAT",
        "COMUNIDAD_AUTONOMA", "SERVICIO_SALUD", "AGENCIA",
    }
    for propuesta in propuestas:
        if propuesta.get("confianza_sede") == "ALTA":
            continue
        administracion = propuesta.get("administracion_detectada", "")
        familia = propuesta.get("familia", "")
        if not administracion or familia in {"AYUNTAMIENTO", "DIPUTACION", "DIPUTACION_PROVINCIAL"}:
            continue
        clave = (administracion, familia)
        pendientes[clave] += propuesta.get("filas_oposiciones", 0)
        if familia in familias_supramunicipales:
            causas[clave] = "ENTIDAD_SUPRAMUNICIPAL"
        else:
            # El catálogo admite únicamente sedes verificadas: una entidad
            # concreta ausente no se interpreta como sede dudosa.
            causas[clave] = "SEDE_NO_ENCONTRADA"
    total = sum(pendientes.values())
    ranking, acumulado = [], 0
    for posicion, ((administracion, familia), filas) in enumerate(pendientes.most_common(), start=1):
        acumulado += filas
        ranking.append({
            "posicion": posicion,
            "Administracion": administracion,
            "familia": familia,
            "filas_oposiciones": filas,
            "filas_acumuladas": acumulado,
            "porcentaje_acumulado": round(100 * acumulado / total, 2) if total else 0,
            "causa": causas[(administracion, familia)],
        })
    umbrales = {}
    for porcentaje in (50, 75, 90, 95):
        objetivo = total * porcentaje / 100
        fila = next((x for x in ranking if x["filas_acumuladas"] >= objetivo), None)
        umbrales[str(porcentaje)] = {
            "administraciones_necesarias": fila["posicion"] if fila else 0,
            "filas_cubiertas": fila["filas_acumuladas"] if fila else 0,
        }
    return {"filas_requieren_sede": total, "ranking_completo": ranking, "umbrales_cobertura": umbrales}


def _cargar_catalogo_completo(ruta_catalogo):
    catalogo = json.loads(Path(ruta_catalogo).read_text(encoding="utf-8"))
    pendientes = set(catalogo.get("fechas_objetivo", [])) - set(catalogo.get("fechas_completadas", []))
    if pendientes or catalogo.get("errores_por_fecha"):
        raise ValueError("El catálogo BOE no está completo o contiene errores de fecha")
    return catalogo


def calcular_propuestas(excel="BOE-oposiciones.xlsx", ruta_catalogo=RUTA_CATALOGO, stream=None):
    """Fuente única de propuestas para dry-run y commit, sin consultar al BOE."""
    catalogo = _cargar_catalogo_completo(ruta_catalogo)
    # Fuerza la validación de los tres catálogos antes de leer o preparar XLSX.
    cargar_capitales_provinciales(); cargar_sedes_administraciones(); cargar_alias_municipios()
    publicaciones, oposiciones = cargar_historicos(excel)
    return propuesta_dry_run(publicaciones, oposiciones, catalogo, stream)


def _resumen_actualizacion(resultado, actualizaciones):
    return {
        "administraciones_a_actualizar": sum("Administración" in x["valores"] for x in actualizaciones),
        "municipios_a_actualizar": sum("Municipio" in x["valores"] for x in actualizaciones),
        "provincias_a_actualizar": sum("Provincia" in x["valores"] for x in actualizaciones),
        "filas_con_cambios": len(actualizaciones),
        "filas_completamente_geolocalizadas": resultado["resumen"]["filas_completamente_geolocalizables"],
        "no_resueltas": resultado["resumen"]["no_resueltas"],
        "ambiguas": resultado["resumen"]["ambiguas"],
    }


def preparar_actualizaciones(excel, resultado):
    """Calcula únicamente cambios autorizados de Oposiciones, en memoria."""
    propuestas = {x["Publicacion_ID"]: x for x in resultado["propuestas_sede"]}
    libro = load_workbook(excel, read_only=True, data_only=False)
    try:
        if "Oposiciones" not in libro.sheetnames:
            raise ValueError("El Excel no contiene la hoja Oposiciones")
        hoja = libro["Oposiciones"]
        cabeceras = [celda.value for celda in next(hoja.iter_rows(min_row=1, max_row=1))]
        obligatorias = {"Publicacion_ID", "Administración", "Municipio", "Provincia", "Version_extractor"}
        if not obligatorias.issubset(cabeceras):
            raise ValueError("Oposiciones no contiene las columnas necesarias para el enriquecimiento")
        columnas = {nombre: indice + 1 for indice, nombre in enumerate(cabeceras)}
        cambios = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            valores = dict(zip(cabeceras, fila))
            if valores.get("Version_extractor") != VERSION_HISTORICA or not es_generica(valores.get("Administración")):
                continue
            propuesta = propuestas.get(str(valores.get("Publicacion_ID") or ""))
            if not propuesta or propuesta.get("confianza") != "ALTA":
                continue
            nuevos = {}
            if valores.get("Administración") != propuesta["administracion_detectada"]:
                nuevos["Administración"] = propuesta["administracion_detectada"]
            if propuesta.get("confianza_sede") == "ALTA":
                municipio, provincia = propuesta.get("Municipio", ""), propuesta.get("Provincia", "")
                if not municipio or not provincia:
                    raise RuntimeError("Una sede ALTA no contiene Municipio y Provincia")
                if valores.get("Municipio") != municipio:
                    nuevos["Municipio"] = municipio
                if valores.get("Provincia") != provincia:
                    nuevos["Provincia"] = provincia
            if nuevos:
                cambios.append({"fila": numero_fila, "valores": nuevos, "columnas": columnas,
                                "Publicacion_ID": valores.get("Publicacion_ID")})
        return cambios
    finally:
        libro.close()


def _crear_backup_enriquecimiento(excel, directorio="backups/enriquecimiento_administraciones"):
    origen = Path(excel); marca = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = Path(directorio) / f"BOE-oposiciones_pre_enriquecimiento_{marca}.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)
    firma_origen = firma_excel(origen)
    shutil.copy2(origen, destino)
    firma_backup = firma_excel(destino)
    if {k: firma_backup[k] for k in ("sha256", "tamano")} != {k: firma_origen[k] for k in ("sha256", "tamano")}:
        raise RuntimeError("El backup no coincide con el Excel de origen")
    return destino, firma_backup


def _barra_actualizacion(total, stream=None):
    salida = stream or sys.stdout
    return tqdm(total=total, desc="Actualizando Oposiciones", dynamic_ncols=True, file=salida,
                disable=not bool(getattr(salida, "isatty", lambda: False)()))


def _validar_temporal(origen, temporal, cambios):
    """Comprueba que el temporal difiere solo en los campos autorizados."""
    esperados = {(x["fila"], columna): valor for x in cambios for columna, valor in x["valores"].items()}
    original = load_workbook(origen, read_only=True, data_only=False)
    candidato = load_workbook(temporal, read_only=True, data_only=False)
    try:
        if original.sheetnames != candidato.sheetnames:
            raise RuntimeError("El temporal no conserva las mismas hojas")
        catalogo = crear_indice_municipios(cargar_catalogo())
        for nombre in original.sheetnames:
            hoja_original, hoja_temporal = original[nombre], candidato[nombre]
            if hoja_original.max_row != hoja_temporal.max_row or hoja_original.max_column != hoja_temporal.max_column:
                raise RuntimeError(f"El temporal cambió la dimensión de {nombre}")
            filas_original = hoja_original.iter_rows(values_only=True)
            filas_temporal = hoja_temporal.iter_rows(values_only=True)
            cabeceras = next(filas_original); cabeceras_temporal = next(filas_temporal)
            if cabeceras != cabeceras_temporal:
                raise RuntimeError(f"El temporal cambió las columnas de {nombre}")
            indices = {campo: indice + 1 for indice, campo in enumerate(cabeceras)}
            filas_modificadas = {x["fila"] for x in cambios} if nombre == "Oposiciones" else set()
            valores_modificados = {}
            for numero_fila, (fila_original, fila_temporal) in enumerate(zip(filas_original, filas_temporal), start=2):
                for indice, (antes, despues) in enumerate(zip(fila_original, fila_temporal), start=1):
                    if antes == despues:
                        continue
                    campo = cabeceras[indice - 1]
                    if nombre != "Oposiciones" or campo not in {"Administración", "Municipio", "Provincia"}:
                        raise RuntimeError(f"Cambio no autorizado en {nombre}.{campo}")
                    if esperados.get((numero_fila, campo)) != despues:
                        raise RuntimeError("El temporal contiene un cambio de Oposiciones no previsto")
                if numero_fila in filas_modificadas:
                    valores_modificados[numero_fila] = dict(zip(cabeceras, fila_temporal))
            if nombre == "Oposiciones":
                for cambio in cambios:
                    fila = cambio["fila"]
                    municipio = valores_modificados[fila]["Municipio"]
                    provincia = valores_modificados[fila]["Provincia"]
                    if bool(municipio) != bool(provincia):
                        raise RuntimeError("Municipio y Provincia deben escribirse siempre como pareja")
                    if municipio and resolver_entidad(f"{municipio} ({provincia})", catalogo)[-1] != "ALTA":
                        raise RuntimeError("La sede escrita no es válida en municipios_oficial.csv")
    finally:
        original.close(); candidato.close()


def _escribir_actualizaciones_atomicas(excel, cambios, stream=None):
    ruta = Path(excel); descriptor, nombre_temporal = tempfile.mkstemp(
        prefix=f".{ruta.stem}-enriquecimiento-", suffix=".xlsx", dir=ruta.resolve().parent)
    os.close(descriptor); temporal = Path(nombre_temporal)
    try:
        shutil.copy2(ruta, temporal)
        libro = load_workbook(temporal)
        hoja = libro["Oposiciones"]
        barra = _barra_actualizacion(len(cambios), stream)
        try:
            for cambio in cambios:
                for campo, valor in cambio["valores"].items():
                    hoja.cell(cambio["fila"], cambio["columnas"][campo]).value = valor
                barra.update(1)
        finally:
            barra.close()
        libro.save(temporal); libro.close()
        print("Validando Excel temporal...", file=stream or sys.stdout)
        _validar_temporal(ruta, temporal, cambios)
        print("Reemplazando Excel...", file=stream or sys.stdout)
        os.replace(temporal, ruta)
    except BaseException:
        if temporal.exists():
            temporal.unlink()
        raise


def aplicar_enriquecimiento(excel="BOE-oposiciones.xlsx", ruta_catalogo=RUTA_CATALOGO,
                            backup_directorio="backups/enriquecimiento_administraciones", stream=None):
    """Aplica una única actualización transaccional, o informa idempotencia."""
    with bloqueo_excel(excel):
        resultado = calcular_propuestas(excel, ruta_catalogo, stream)
        cambios = preparar_actualizaciones(excel, resultado)
        resumen = _resumen_actualizacion(resultado, cambios)
        if not cambios:
            return {"ya_aplicado": True, **resumen}
        print("Creando backup...", file=stream or sys.stdout)
        backup, firma_backup = _crear_backup_enriquecimiento(excel, backup_directorio)
        print("Escribiendo Excel temporal...", file=stream or sys.stdout)
        _escribir_actualizaciones_atomicas(excel, cambios, stream)
        return {"ya_aplicado": False, "backup": str(backup), "backup_firma": firma_backup, **resumen}


def ejecutar_dry_run(excel="BOE-oposiciones.xlsx", ruta_catalogo=RUTA_CATALOGO, obtener=None, stream=None):
    """Compatibilidad pública: dry-run local sin descargar ni alterar catálogo."""
    return calcular_propuestas(excel, ruta_catalogo, stream)


def escribir_diagnostico_sedes(resultado, directorio="informes/diagnostico_administraciones_historicas"):
    if "sedes" not in resultado: return
    destino = Path(directorio); destino.mkdir(parents=True, exist_ok=True)
    (destino / "diagnostico_sedes_administrativas.json").write_text(json.dumps(resultado, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lineas = ["# Diagnóstico de sedes administrativas", "", "## Resumen", ""]
    lineas += [f"- {k}: {v}" for k, v in resultado["resumen"].items()]
    lineas += ["", "## 100 administraciones sin sede con mayor impacto", "", "| Administración | Familia | Filas |", "|---|---|---:|"]
    lineas += [f"| {x['Administracion']} | {x['familia']} | {x['filas_oposiciones']} |" for x in resultado["no_resueltas_sedes"]]
    pendientes = resultado.get("pendientes_sede", {})
    if pendientes:
        lineas += ["", "## Cobertura acumulada de sedes pendientes", "",
                  "| Objetivo | Administraciones necesarias | Filas cubiertas |", "|---:|---:|---:|"]
        lineas += [f"| {objetivo}% | {datos['administraciones_necesarias']} | {datos['filas_cubiertas']} |"
                   for objetivo, datos in pendientes["umbrales_cobertura"].items()]
        lineas += ["", "## Ranking completo de entidades que requieren sede", "",
                  "| # | Administración | Familia | Filas | Acumulado | % acumulado | Causa |",
                  "|---:|---|---|---:|---:|---:|---|"]
        lineas += [f"| {x['posicion']} | {x['Administracion']} | {x['familia']} | {x['filas_oposiciones']} | "
                   f"{x['filas_acumuladas']} | {x['porcentaje_acumulado']}% | {x['causa']} |"
                   for x in pendientes["ranking_completo"]]
    (destino / "diagnostico_sedes_administrativas.md").write_text("\n".join(lineas) + "\n", encoding="utf-8")


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    modo = parser.add_mutually_exclusive_group(required=True)
    modo.add_argument("--dry-run", action="store_true")
    modo.add_argument("--aplicar", action="store_true")
    parser.add_argument("--excel", default="BOE-oposiciones.xlsx")
    parser.add_argument("--catalogo", default=RUTA_CATALOGO)
    args = parser.parse_args(argv)
    antes = firma_excel(args.excel)
    inicio = time.monotonic()
    if args.dry_run:
        print("DRY-RUN: BOE-oposiciones.xlsx NO será modificado.")
        resultado = ejecutar_dry_run(args.excel, args.catalogo)
        resultado["actualizacion"] = {
            **resultado.get("prevision_actualizacion", {}),
            "filas_con_cambios": resultado.get("prevision_actualizacion", {}).get("administraciones_a_actualizar", 0),
            "filas_completamente_geolocalizadas": resultado["resumen"]["filas_completamente_geolocalizables"],
            "no_resueltas": resultado["resumen"]["no_resueltas"],
            "ambiguas": resultado["resumen"]["ambiguas"],
        }
    else:
        resultado = aplicar_enriquecimiento(args.excel, args.catalogo)
    resultado["tiempo_segundos"] = round(time.monotonic() - inicio, 2)
    despues = firma_excel(args.excel)
    if args.dry_run and antes != despues:
        raise RuntimeError("El dry-run modificó el Excel")
    if args.dry_run:
        escribir_diagnostico_sedes(resultado)
        salida = {**resultado["resumen"], **resultado["actualizacion"], "tiempo_segundos": resultado["tiempo_segundos"]}
    else:
        salida = resultado
    print(json.dumps(salida, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
