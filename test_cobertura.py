from datetime import datetime

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import preparar_archivo_datos
from cobertura import (
    COLUMNAS_COBERTURA,
    cobertura_indice_reutilizable,
    normalizar_cobertura,
    puede_reutilizar_cobertura,
    registrar_cobertura,
)


def test_crea_registro_con_columnas_en_orden_y_fecha_estable():
    resultado = registrar_cobertura(
        pd.DataFrame(),
        "2026/08/20",
        "consultado",
        8,
        datetime(2026, 8, 22, 12, 34, 56),
    )

    assert resultado.columns.tolist() == COLUMNAS_COBERTURA
    assert resultado.iloc[0].to_dict() == {
        "Fecha": "2026-08-20",
        "Estado": "consultado",
        "Version_extractor": "1",
        "Fecha_ultima_consulta": "2026-08-22 12:34:56",
        "Numero_publicaciones": 8,
    }


@pytest.mark.filterwarnings("error::FutureWarning")
@pytest.mark.parametrize(
    "inicial",
    [
        pd.DataFrame(),
        pd.DataFrame(
            [
                {
                    "Fecha": "2026-08-20",
                    "Estado": "error",
                    "Version_extractor": float("nan"),
                    "Fecha_ultima_consulta": "2026-08-21 10:00:00",
                    "Numero_publicaciones": float("nan"),
                }
            ]
        ),
    ],
)
def test_version_textual_no_emite_futurewarning_y_sigue_siendo_reutilizable(
    inicial,
):
    resultado = registrar_cobertura(
        inicial,
        "2026-08-20",
        "consultado",
        0,
        momento="2026-08-22 12:00:00",
        version_actual="1",
    )

    assert resultado.loc[0, "Version_extractor"] == "1"
    assert resultado["Version_extractor"].dtype == object
    assert puede_reutilizar_cobertura(
        "2026-08-20", resultado, pd.DataFrame(), pd.DataFrame()
    )


def test_actualiza_fecha_existente_sin_duplicarla_y_error_se_sustituye():
    inicial = registrar_cobertura(
        pd.DataFrame(), "2026-08-20", "error", momento="primer intento"
    )

    resultado = registrar_cobertura(
        inicial, "2026-08-20", "consultado", 3, momento="segundo intento"
    )

    assert len(resultado) == 1
    assert resultado.iloc[0]["Estado"] == "consultado"
    assert resultado.iloc[0]["Numero_publicaciones"] == 3
    assert resultado.iloc[0]["Fecha_ultima_consulta"] == "segundo intento"


def test_error_posterior_preserva_cobertura_correcta_y_actualiza_intento():
    inicial = registrar_cobertura(
        pd.DataFrame(), "2026-08-20", "sin_edicion", 0, momento="correcta"
    )

    resultado = registrar_cobertura(
        inicial, "2026-08-20", "error", momento="fallida"
    )

    assert len(resultado) == 1
    assert resultado.iloc[0]["Estado"] == "sin_edicion"
    assert resultado.iloc[0]["Version_extractor"] == "1"
    assert resultado.iloc[0]["Numero_publicaciones"] == 0
    assert resultado.iloc[0]["Fecha_ultima_consulta"] == "fallida"


def test_registro_y_normalizacion_no_modifican_dataframe_original():
    original = pd.DataFrame(
        [{"Fecha": "2026-08-20", "Estado": "error"}]
    )
    copia = original.copy(deep=True)

    registrar_cobertura(original, "2026-08-20", "consultado", 1)
    normalizar_cobertura(original)

    pd.testing.assert_frame_equal(original, copia)


def test_libro_sin_cobertura_la_crea_vacia_sin_reconstruir_historico(
    monkeypatch, tmp_path
):
    ruta = tmp_path / "BOE-oposiciones.xlsx"
    with pd.ExcelWriter(ruta) as writer:
        pd.DataFrame({"Código": ["histórico"]}).to_excel(
            writer, sheet_name="Búsquedas", index=False
        )
        pd.DataFrame({"Puesto": ["Ingeniero"]}).to_excel(
            writer, sheet_name="Oposiciones", index=False
        )
        pd.DataFrame(columns=["Fecha", "Tipo de error", "Enlace Web"]).to_excel(
            writer, sheet_name="Log-errores", index=False
        )
    monkeypatch.chdir(tmp_path)

    hojas = preparar_archivo_datos.preparar_excel_y_dataframes()

    assert hojas["Cobertura"].empty
    assert hojas["Cobertura"].columns.tolist() == COLUMNAS_COBERTURA


def test_escritura_atomica_incluye_cobertura_y_mantiene_busquedas_oculta(
    monkeypatch, tmp_path
):
    ruta = tmp_path / "BOE-oposiciones.xlsx"
    libro = Workbook()
    libro.active.title = "Búsquedas"
    libro.create_sheet("Oposiciones")
    libro.create_sheet("Log-errores")
    libro.save(ruta)
    monkeypatch.chdir(tmp_path)
    cobertura = registrar_cobertura(
        pd.DataFrame(), "2026-08-20", "consultado", 2
    )

    preparar_archivo_datos.guardar_excel(
        pd.DataFrame({"Puesto": ["Ingeniero"]}),
        pd.DataFrame({"Código": ["codigo"]}),
        pd.DataFrame(columns=["Fecha", "Tipo de error", "Enlace Web"]),
        pd.DataFrame(),
        cobertura,
    )

    comprobacion = load_workbook(ruta, read_only=True)
    assert "Cobertura" in comprobacion.sheetnames
    assert comprobacion["Búsquedas"].sheet_state == "hidden"
    comprobacion.close()
    guardada = pd.read_excel(ruta, sheet_name="Cobertura")
    assert guardada.columns.tolist() == COLUMNAS_COBERTURA
    assert guardada.loc[0, "Numero_publicaciones"] == 2


def _cobertura(estado="consultado", version="1", numero=0, fecha="2026-08-10"):
    return pd.DataFrame(
        [
            {
                "Fecha": fecha,
                "Estado": estado,
                "Version_extractor": version,
                "Fecha_ultima_consulta": "2026-08-22 12:00:00",
                "Numero_publicaciones": numero,
            }
        ]
    )


def _publicaciones(cantidad, version="1", estado="sin_coincidencias"):
    return pd.DataFrame(
        [
            {
                "Publicacion_ID": f"BOE-A-2026-{1000 + indice}",
                "Fecha_BOE": "10 de agosto de 2026",
                "Version_extractor": version,
                "Estado_analisis": estado,
                "Coincidencias": 0 if estado == "sin_coincidencias" else 1,
            }
            for indice in range(cantidad)
        ]
    )


@pytest.mark.parametrize(
    ("cobertura", "esperado"),
    [
        (pd.DataFrame(), False),
        (_cobertura("error"), False),
        (_cobertura(version="legacy"), True),
        (_cobertura(version="1"), True),
        (_cobertura("sin_edicion", "2", 0), True),
        (_cobertura("consultado", "2", 0), True),
        (_cobertura("sin_edicion", "1", 1), False),
    ],
)
def test_decision_basica_de_reutilizacion(cobertura, esperado):
    assert (
        puede_reutilizar_cobertura(
            "2026/08/10",
            cobertura,
            pd.DataFrame(),
            pd.DataFrame(),
            version_actual="2",
        )
        is esperado
    )


@pytest.mark.parametrize(("locales", "esperado"), [(8, True), (7, False), (9, False)])
def test_numero_de_publicaciones_locales_debe_coincidir(locales, esperado):
    assert (
        puede_reutilizar_cobertura(
            "2026-08-10",
            _cobertura(numero=8),
            _publicaciones(locales),
            pd.DataFrame(),
        )
        is esperado
    )


@pytest.mark.parametrize("version", ["historico-experimental-2004", "legacy", "0", ""])
def test_version_de_analisis_no_invalida_cobertura_indice(version):
    assert puede_reutilizar_cobertura(
        "2026-08-10",
        _cobertura(numero=1),
        _publicaciones(1, version=version),
        pd.DataFrame(),
    )


def test_cobertura_indice_no_exige_oposiciones_ya_que_es_analisis_separado():
    publicaciones = _publicaciones(1, estado="con_coincidencias")

    assert puede_reutilizar_cobertura(
        "2026-08-10", _cobertura(numero=1), publicaciones, pd.DataFrame()
    )


def test_historico_sin_edicion_con_publicaciones_es_contradictorio():
    publicaciones = _publicaciones(1, version="historico-experimental-2004")
    assert not cobertura_indice_reutilizable(
        "2026-08-10", _cobertura("sin_edicion", "historico-experimental-2004", 0), publicaciones
    )


def test_version_futura_no_invalida_cobertura_y_no_etiqueta_publicacion():
    publicaciones = _publicaciones(1, version="historico-experimental-2004")
    assert cobertura_indice_reutilizable(
        "2026-08-10", _cobertura("consultado", "historico-experimental-2004", 1), publicaciones
    )
    assert publicaciones.loc[0, "Version_extractor"] == "historico-experimental-2004"


def test_sin_coincidencias_no_exige_fila_en_oposiciones():
    assert puede_reutilizar_cobertura(
        "2026-08-10",
        _cobertura(numero=1),
        _publicaciones(1),
        pd.DataFrame(),
    )


@pytest.mark.parametrize("version", ["1", "2"])
def test_publicacion_con_version_actual_o_posterior_es_compatible(version):
    assert puede_reutilizar_cobertura(
        "2026-08-10",
        _cobertura(numero=1),
        _publicaciones(1, version=version),
        pd.DataFrame(),
    )


def test_publicacion_sin_identificador_valido_impide_reutilizar():
    publicaciones = _publicaciones(1)
    publicaciones.loc[0, "Publicacion_ID"] = "identificador-invalido"

    assert not puede_reutilizar_cobertura(
        "2026-08-10", _cobertura(numero=1), publicaciones, pd.DataFrame()
    )


def test_fechas_equivalentes_con_tipos_distintos_se_reutilizan():
    cobertura = _cobertura(numero=1, fecha=pd.Timestamp("2026-08-10"))
    publicaciones = _publicaciones(1)
    publicaciones.loc[0, "Fecha_BOE"] = datetime(2026, 8, 10)

    assert puede_reutilizar_cobertura(
        "10/08/2026", cobertura, publicaciones, pd.DataFrame()
    )


def test_fecha_invalida_no_se_reutiliza():
    assert not puede_reutilizar_cobertura(
        "fecha inválida", _cobertura(), pd.DataFrame(), pd.DataFrame()
    )


def test_version_posterior_de_cobertura_es_compatible():
    assert puede_reutilizar_cobertura(
        "2026-08-10",
        _cobertura("consultado", "2", 0),
        pd.DataFrame(),
        pd.DataFrame(),
    )


@pytest.mark.parametrize("numero", [pd.NA, "invalido", 1.5, -1])
def test_numero_publicaciones_invalido_impide_reutilizar(numero):
    assert not puede_reutilizar_cobertura(
        "2026-08-10",
        _cobertura("consultado", "1", numero),
        pd.DataFrame(),
        pd.DataFrame(),
    )


def test_reutilizacion_no_modifica_ningun_dataframe():
    cobertura = _cobertura(numero=1)
    publicaciones = _publicaciones(1)
    oposiciones = pd.DataFrame()
    copias = [dataframe.copy(deep=True) for dataframe in (cobertura, publicaciones, oposiciones)]

    assert puede_reutilizar_cobertura(
        "2026-08-10", cobertura, publicaciones, oposiciones
    )

    for dataframe, copia in zip((cobertura, publicaciones, oposiciones), copias):
        pd.testing.assert_frame_equal(dataframe, copia)


def test_incoherencia_historica_verificada_es_reutilizable_sin_fingir_coherencia():
    publicaciones = _publicaciones(1)

    assert cobertura_indice_reutilizable(
        "2026-08-10",
        _cobertura("incoherencia_historica_verificada", numero=0),
        publicaciones,
    )
    assert not cobertura_indice_reutilizable(
        "2026-08-10", _cobertura("consultado", numero=0), publicaciones
    )
