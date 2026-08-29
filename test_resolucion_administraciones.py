from resolucion_administraciones import (
    ResolutorAdministraciones, enriquecer_convocatorias,
    resolver_administracion,
)
import pandas as pd
from preparar_archivo_datos import combinar_dataframes
from publicaciones import crear_registro_publicacion
from openpyxl import Workbook, load_workbook


def test_titulo_ambiguo_y_departamento_vacio_no_inventan_administracion():
    ambigua = resolver_administracion(
        "Resolución del Ayuntamiento de A, referente a la convocatoria del Consorcio B."
    )
    vacia = resolver_administracion("", "")
    assert (ambigua.estado, ambigua.confianza) == ("AMBIGUA", "AMBIGUA")
    assert (vacia.estado, vacia.administracion) == ("NO_RESUELTA", "")


def test_enriquece_todas_las_filas_genericas_y_conserva_la_especifica():
    resolutor = ResolutorAdministraciones()
    filas, administracion = enriquecer_convocatorias(
        [
            {"Puesto": "A", "Administración": "Administración Local"},
            {"Puesto": "B", "Administración": "Ayuntamiento específico"},
            {"Puesto": "C", "Administración": "--"},
        ],
        {"titulo": "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria.", "departamento": "Administración Local"},
        resolutor,
    )
    assert administracion.administracion == "Ayuntamiento de Ciudad Real"
    assert [x["Administración"] for x in filas] == [
        "Ayuntamiento de Ciudad Real", "Ayuntamiento específico", "Ayuntamiento de Ciudad Real"
    ]
    assert [(x.get("Municipio"), x.get("Provincia")) for x in (filas[0], filas[2])] == [
        ("Ciudad Real", "Ciudad Real"), ("Ciudad Real", "Ciudad Real")
    ]
    assert "Municipio" not in filas[1]


def test_alias_y_diputacion_reales_son_deterministas_y_sin_fuzzy():
    resolutor = ResolutorAdministraciones()
    alias = resolutor.resolver_sede("Ayuntamiento de Santa María de Guía (Las Palmas)", "AYUNTAMIENTO")
    diputacion = resolutor.resolver_sede("Diputación Provincial de Castellón", "DIPUTACION_PROVINCIAL")
    inexistente = resolutor.resolver_sede("Ayuntamiento de Santa María de Gu", "AYUNTAMIENTO")
    assert (alias.municipio, alias.codigo_ine, alias.confianza) == ("Santa María de Guía de Gran Canaria", "35023", "ALTA")
    assert (diputacion.municipio, diputacion.confianza) == ("Castelló de la Plana/Castellón de la Plana", "ALTA")
    assert inexistente.confianza == "NO_RESUELTA"


def test_enriquecimiento_previo_a_deduplicacion_conserva_campos_funcionales():
    original = {"Puesto": "Auxiliar", "Num_plazas": 1, "Fecha_boe": "1 de enero de 2026",
                "Enlace": "https://x?id=BOE-A-2026-1", "Turno": "--", "Sistema": "--",
                "Escala": "--", "Subescala": "--", "Clase": "--", "Administración": "Administración Local"}
    filas, _ = enriquecer_convocatorias([original], {
        "titulo": "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria.", "departamento": ""
    })
    especifica = {**original, "Administración": "Ayuntamiento de Ciudad Real", "Municipio": "Ciudad Real", "Provincia": "Ciudad Real"}
    combinado, _ = combinar_dataframes(
        {k: [v] for k, v in filas[0].items()}, {"Código": []}, pd.DataFrame([especifica]), pd.DataFrame({"Código": []})
    )
    assert len(combinado) == 1
    assert (filas[0]["Puesto"], filas[0]["Num_plazas"], filas[0]["Fecha_boe"], filas[0]["Enlace"]) == (
        original["Puesto"], original["Num_plazas"], original["Fecha_boe"], original["Enlace"]
    )


def test_administracion_especifica_no_se_sobrescribe_ni_se_inventan_sedes():
    filas, _ = enriquecer_convocatorias(
        [{"Puesto": "P", "Num_plazas": 1, "Administración": "Entidad específica", "Municipio": "M", "Provincia": "P"}],
        {"titulo": "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria."},
    )
    assert filas == [{"Puesto": "P", "Num_plazas": 1, "Administración": "Entidad específica", "Municipio": "M", "Provincia": "P"}]


def test_end_to_end_sumario_a_filas_y_trazabilidad_comparten_motor():
    titulo = "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria."
    filas, resolucion = enriquecer_convocatorias(
        [{"Puesto": "Auxiliar", "Num_plazas": 1, "Administración": "Administración Local"}],
        {"titulo": titulo, "departamento": "Administración Local"},
    )
    publicacion = crear_registro_publicacion(
        "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-1", "1 de enero de 2026", "", 1,
        titulo_sumario=titulo, departamento_boe="Administración Local",
    )
    assert (filas[0]["Administración"], filas[0]["Municipio"], filas[0]["Provincia"]) == (
        publicacion["Administracion_resuelta"], "Ciudad Real", "Ciudad Real"
    )
    assert (resolucion.familia, publicacion["Familia_administrativa"], publicacion["Version_resolucion"]) == (
        "AYUNTAMIENTO", "AYUNTAMIENTO", "1"
    )


def test_end_to_end_excel_temporal_con_segunda_convocatoria_especifica(tmp_path):
    titulo = "Resolución del Ayuntamiento de Guadalajara, referente a la convocatoria."
    filas, _ = enriquecer_convocatorias([
        {"Puesto": "Auxiliar", "Num_plazas": 1, "Administración": "Administración Local", "Municipio": "", "Provincia": ""},
        {"Puesto": "Técnico", "Num_plazas": 1, "Administración": "Entidad específica", "Municipio": "", "Provincia": ""},
    ], {"titulo": titulo, "departamento": "Administración Local"})
    publicacion = crear_registro_publicacion("https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-2",
        "1 de enero de 2026", "", 2, titulo_sumario=titulo, departamento_boe="Administración Local")
    ruta = tmp_path / "resultado.xlsx"; libro = Workbook(); oposiciones = libro.active; oposiciones.title = "Oposiciones"
    oposiciones.append(list(filas[0])); [oposiciones.append([fila.get(k) for k in filas[0]]) for fila in filas]
    publicaciones = libro.create_sheet("Publicaciones"); publicaciones.append(list(publicacion)); publicaciones.append(list(publicacion.values()))
    libro.save(ruta); libro.close()
    comprobacion = load_workbook(ruta, read_only=True, data_only=True)
    try:
        opos = list(comprobacion["Oposiciones"].iter_rows(values_only=True)); pubs = list(comprobacion["Publicaciones"].iter_rows(values_only=True))
        assert opos[1][2:5] == ("Ayuntamiento de Guadalajara", "Guadalajara", "Guadalajara")
        assert opos[2][2] == "Entidad específica"
        assert dict(zip(pubs[0], pubs[1]))["Estado_resolucion"] == "RESUELTA"
    finally:
        comprobacion.close()
