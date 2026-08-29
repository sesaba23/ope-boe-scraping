import json
from io import StringIO

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from enriquecer_administraciones_historicas import (
    construir_catalogo, extraer_administraciones_titulo, fechas_historicas, propuesta_dry_run, ProgresoFechas, resolver_sedes,
    analizar_alias_municipios, analizar_pendientes_sede,
    _crear_backup_enriquecimiento, _escribir_actualizaciones_atomicas,
    _validar_temporal, aplicar_enriquecimiento, preparar_actualizaciones,
)


def _excel_prueba(ruta):
    libro = Workbook()
    libro.active.title = "Búsquedas"
    libro["Búsquedas"].append(["Código", "Texto"]); libro["Búsquedas"].append(["x", "sin cambios"])
    oposiciones = libro.create_sheet("Oposiciones")
    oposiciones.append(["Publicacion_ID", "Administración", "Municipio", "Provincia", "Version_extractor", "Puesto", "Num_plazas", "Turno"])
    oposiciones.append(["BOE-A-1", "Administración Local", None, None, "historico-experimental-2004", "Puesto 1", 1, "--"])
    oposiciones.append(["BOE-A-2", "Administración Local", None, None, "historico-experimental-2004", "Puesto 2", 2, "--"])
    oposiciones.append(["BOE-A-3", "Administración Local", None, None, "historico-experimental-2004", "Puesto 3", 3, "--"])
    for nombre in ("Log-errores", "Publicaciones", "Cobertura"):
        hoja = libro.create_sheet(nombre); hoja.append(["columna"]); hoja.append([nombre])
    libro.save(ruta)


def _resultado_prueba():
    return {"resumen": {"filas_completamente_geolocalizables": 1, "no_resueltas": 1, "ambiguas": 1}, "propuestas_sede": [
        {"Publicacion_ID": "BOE-A-1", "confianza": "ALTA", "administracion_detectada": "Universidad de Salamanca", "confianza_sede": "NO_RESUELTA", "Municipio": "", "Provincia": ""},
        {"Publicacion_ID": "BOE-A-2", "confianza": "ALTA", "administracion_detectada": "Ayuntamiento de Ciudad Real", "confianza_sede": "ALTA", "Municipio": "Ciudad Real", "Provincia": "Ciudad Real"},
        {"Publicacion_ID": "BOE-A-3", "confianza": "AMBIGUA", "administracion_detectada": "", "confianza_sede": "NO_RESUELTA", "Municipio": "", "Provincia": ""},
    ]}


@pytest.mark.parametrize(("titulo", "familia", "administracion"), [
    ("Resolución de 2 de abril de 2004, de la Diputación Provincial de Albacete, referente a la convocatoria.", "DIPUTACION_PROVINCIAL", "Diputación Provincial de Albacete"),
    ("Resolución de 17 de mayo de 2005, del Ayuntamiento de Benalmádena (Málaga), referente a la convocatoria.", "AYUNTAMIENTO", "Ayuntamiento de Benalmádena (Málaga)"),
    ("Resolución de la Universidad de Salamanca, referente a la convocatoria.", "UNIVERSIDAD", "Universidad de Salamanca"),
    ("Resolución de 12 de febrero de 2008, del Consorcio Hospitalario Provincial de Castellón, referente a la convocatoria.", "CONSORCIO", "Consorcio Hospitalario Provincial de Castellón"),
])
def test_extrae_familias_reales_del_diagnostico(titulo, familia, administracion):
    resultado = extraer_administraciones_titulo(titulo)
    assert (resultado["familia"], resultado["administracion_detectada"], resultado["confianza"]) == (familia, administracion, "ALTA")


def test_titulo_no_resoluble_y_multiple_administracion_no_se_proponen():
    assert extraer_administraciones_titulo("Resolución sin entidad convocante")["confianza"] == "NO_RESUELTA"
    assert extraer_administraciones_titulo("Resolución del Ayuntamiento de A, referente a la convocatoria del Consorcio B.")["confianza"] == "AMBIGUA"


def test_catalogo_reanuda_y_no_repite_fecha(tmp_path):
    publicaciones = pd.DataFrame([{"Publicacion_ID": "BOE-A-2004-1", "Fecha_BOE": "2004-01-01"}])
    llamadas = []
    def obtener(fecha):
        llamadas.append(fecha)
        return {"estado": "OK", "sumario": {"diario": [{"seccion": [{"codigo": "2B", "departamento": [{"nombre": "Administración Local", "item": [{"identificador": "BOE-A-2004-1", "titulo": "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria.", "url_html": "h", "url_xml": "x"}]}]}]}]}}
    ruta = tmp_path / "catalogo.json"
    primero = construir_catalogo(publicaciones, ruta, obtener)
    segundo = construir_catalogo(publicaciones, ruta, obtener)
    assert llamadas == ["2004-01-01"]
    assert primero["publicaciones"] == segundo["publicaciones"]
    assert segundo["publicacion_ids_objetivo"] == ["BOE-A-2004-1"]
    assert json.loads(ruta.read_text())["fechas_completadas"] == ["2004-01-01"]


def test_fechas_historicas_admite_fecha_boe_en_castellano_e_iso():
    publicaciones = pd.DataFrame([
        {"Fecha_BOE": "15 de abril de 2024"}, {"Fecha_BOE": "2004-01-01"},
    ])
    assert fechas_historicas(publicaciones) == ["2004-01-01", "2024-04-15"]


def test_fallo_api_persiste_y_es_reintentable(tmp_path):
    publicaciones = pd.DataFrame([{"Publicacion_ID": "BOE-A-2004-1", "Fecha_BOE": "2004-01-01"}])
    ruta = tmp_path / "catalogo.json"
    construir_catalogo(publicaciones, ruta, lambda _: (_ for _ in ()).throw(RuntimeError("temporal")))
    datos = json.loads(ruta.read_text())
    assert "2004-01-01" in datos["errores_por_fecha"] and not datos["fechas_completadas"]
    def exito(_):
        return {"estado": "OK", "sumario": {"diario": [{"seccion": []}]}}
    recuperado = construir_catalogo(publicaciones, ruta, exito)
    assert recuperado["fechas_completadas"] == ["2004-01-01"]


def test_propuesta_propaga_varias_filas_y_no_escribe_excel():
    publicaciones = pd.DataFrame([{"Publicacion_ID": "BOE-A-2004-1", "Fecha_BOE": "2004-01-01"}])
    oposiciones = pd.DataFrame([{"Publicacion_ID": "BOE-A-2004-1", "Administración": "Administración Local"}] * 2)
    catalogo = {"fechas_objetivo": ["2004-01-01"], "fechas_completadas": ["2004-01-01"], "errores_por_fecha": {},
                "publicaciones": {"BOE-A-2004-1": {"titulo": "Resolución del Ayuntamiento de Ciudad Real, referente a la convocatoria.", "departamento": "Administración Local"}}}
    resultado = propuesta_dry_run(publicaciones, oposiciones, catalogo)
    assert resultado["resumen"]["filas_potencialmente_corregibles"] == 2
    assert resultado["propuestas"][0]["administracion_detectada"] == "Ayuntamiento de Ciudad Real"


def test_progreso_no_tty_informa_sin_caracteres_de_control():
    salida = StringIO()
    progreso = ProgresoFechas(2, acumuladas=3, publicaciones=4, stream=salida)
    progreso.actualizar(); progreso.actualizar(); progreso.cerrar()
    assert "Sumarios pendientes: 0/2" in salida.getvalue()
    assert "\x1b" not in salida.getvalue()


def test_resuelve_ayuntamiento_y_no_inventa_sede_universitaria():
    propuestas = [
        {"administracion_detectada": "Ayuntamiento de Ciudad Real", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 2},
        {"administracion_detectada": "Universidad de Salamanca", "familia": "UNIVERSIDAD", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    resultado = resolver_sedes(propuestas, StringIO())
    assert resultado["sedes"][0]["Municipio"] == "Ciudad Real"
    assert resultado["sedes"][1]["confianza"] == "NO_RESUELTA"


def test_diputaciones_usam_capitales_catalogados_y_bilingues():
    propuestas = [
        {"administracion_detectada": "Diputación Provincial de Albacete", "familia": "DIPUTACION_PROVINCIAL", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Diputación Provincial de Castellón", "familia": "DIPUTACION_PROVINCIAL", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert sedes[0]["Municipio"] == "Albacete"
    assert sedes[1]["Municipio"] == "Castelló de la Plana/Castellón de la Plana"
    assert "CATALOGO_CAPITALES" in sedes[1]["metodo_resolucion"]


def test_las_52_capitales_provinciales_se_validan_con_catalogo_oficial():
    from enriquecer_administraciones_historicas import cargar_capitales_provinciales
    from diagnostico_administraciones_historicas import cargar_catalogo, crear_indice_municipios, resolver_entidad
    indice = crear_indice_municipios(cargar_catalogo())
    capitales = {id(x): x for valores in cargar_capitales_provinciales().values() for x in valores}.values()
    assert len(list(capitales)) == 52
    assert all(resolver_entidad(f"{x['Municipio_catalogo']} ({x['Provincia']})", indice)[-1] == "ALTA" for x in capitales)


def test_catalogo_sedes_valida_cabildo_consejo_y_consorcio_con_municipios():
    propuestas = [
        {"administracion_detectada": "Cabildo Insular de Tenerife", "familia": "CABILDO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Consejo Insular de Mallorca", "familia": "CONSEJO_INSULAR", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Consorcio Hospitalario Provincial de Castellón", "familia": "CONSORCIO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert all(x["confianza"] == "ALTA" for x in sedes)
    assert all("CATALOGO_SEDES_VALIDADO" in x["metodo_resolucion"] for x in sedes)


def test_ayuntamiento_oficial_y_administracion_no_catalogada_no_se_aproximan():
    propuestas = [
        {"administracion_detectada": "Ayuntamiento de Torredelcampo", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Consorcio inexistente", "familia": "CONSORCIO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert sedes[0]["confianza"] == "ALTA"
    assert sedes[1]["confianza"] == "NO_RESUELTA"


def test_sede_catalogada_debe_existir_en_municipios(tmp_path):
    sedes = tmp_path / "sedes.csv"
    sedes.write_text(
        "Administracion;Familia;Municipio;Provincia;Fuente;Confianza\n"
        "Consorcio de prueba;CONSORCIO;Municipio inexistente;Madrid;https://entidad.test/;ALTA\n",
        encoding="utf-8",
    )
    propuestas = [{"administracion_detectada": "Consorcio de prueba", "familia": "CONSORCIO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1}]
    resultado = resolver_sedes(propuestas, StringIO(), ruta_sedes=sedes)
    assert resultado["sedes"][0]["confianza"] == "NO_RESUELTA"
    assert resultado["sedes"][0]["metodo_resolucion"] == "CATALOGO_SEDES_MUNICIPIO_NO_VALIDADO"


def test_catalogo_exige_url_institucional_y_confianza_controlada(tmp_path):
    from enriquecer_administraciones_historicas import cargar_sedes_administraciones
    ruta = tmp_path / "sedes.csv"
    ruta.write_text(
        "Administracion;Familia;Municipio;Provincia;Fuente;Confianza\n"
        "Entidad;CONSORCIO;Bilbao;Bizkaia;sin-url;BAJA\n", encoding="utf-8"
    )
    with pytest.raises(ValueError, match="URL institucional HTTPS"):
        cargar_sedes_administraciones(ruta)


def test_variantes_documentales_cabildo_y_mancomunidad_comparten_sede_catalogada():
    propuestas = [
        {"administracion_detectada": "Cabildo Insular de Gran Canaria (Las Palmas)", "familia": "CABILDO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Cabildo Insular de Gran Canaria, Turismo de Gran Canaria (Las Palmas)", "familia": "CABILDO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Mancomunidad de Islantilla (Huelva)", "familia": "MANCOMUNIDAD", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert [x["Municipio"] for x in sedes] == ["Palmas de Gran Canaria, Las", "Palmas de Gran Canaria, Las", "Isla Cristina"]
    assert all(x["confianza"] == "ALTA" for x in sedes)


def test_ranking_pendiente_calcula_acumulados_y_umbrales_sin_inferir_sedes():
    propuestas = [
        {"administracion_detectada": "Consorcio A", "familia": "CONSORCIO", "confianza_sede": "NO_RESUELTA", "filas_oposiciones": 6},
        {"administracion_detectada": "Instituto B", "familia": "INSTITUTO", "confianza_sede": "NO_RESUELTA", "filas_oposiciones": 3},
        {"administracion_detectada": "Ayuntamiento de C", "familia": "AYUNTAMIENTO", "confianza_sede": "NO_RESUELTA", "filas_oposiciones": 9},
    ]
    resultado = analizar_pendientes_sede(propuestas)
    assert resultado["filas_requieren_sede"] == 9
    assert [x["Administracion"] for x in resultado["ranking_completo"]] == ["Consorcio A", "Instituto B"]
    assert resultado["umbrales_cobertura"]["50"]["administraciones_necesarias"] == 1
    assert resultado["ranking_completo"][0]["causa"] == "SEDE_NO_ENCONTRADA"


def test_alias_municipales_reales_resuelven_historico_bilingue_y_con_provincia():
    propuestas = [
        {"administracion_detectada": "Ayuntamiento de Santa María de Guía (Las Palmas)", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Ayuntamiento de Villa de Garafía (Santa Cruz de Tenerife)", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Ayuntamiento de Vilanova y la Geltrú (Barcelona)", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert [(x["Municipio"], x["Provincia"]) for x in sedes] == [
        ("Santa María de Guía de Gran Canaria", "Las Palmas"),
        ("Garafía", "Santa Cruz de Tenerife"),
        ("Vilanova i la Geltrú", "Barcelona"),
    ]
    assert all("CATALOGO_ALIAS_MUNICIPIOS" in x["metodo_resolucion"] for x in sedes)


def test_alias_no_sustituye_coincidencia_oficial_ni_resuelve_ambiguedad():
    propuestas = [
        {"administracion_detectada": "Ayuntamiento de Ciudad Real", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Ayuntamiento de Granada", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Ayuntamiento de Santa María de Gu", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert sedes[0]["metodo_resolucion"].startswith("AYUNTAMIENTO_CATALOGO_MUNICIPIOS_EXACTO")
    assert sedes[1]["confianza"] == "ALTA"
    assert sedes[2]["confianza"] == "NO_RESUELTA"


def test_catalogo_alias_exige_codigo_ine_fuente_y_provincia_validos(tmp_path):
    from enriquecer_administraciones_historicas import cargar_alias_municipios
    ruta = tmp_path / "alias.csv"
    ruta.write_text(
        "Alias;Provincia;Municipio_oficial;Codigo_INE;Fuente;Confianza\n"
        "Alias;Madrid;Municipio inventado;99999;sin-url;ALTA\n", encoding="utf-8"
    )
    with pytest.raises(ValueError):
        cargar_alias_municipios(ruta)


def test_coincidencia_oficial_tiene_prioridad_sobre_alias(tmp_path):
    ruta = tmp_path / "alias.csv"
    ruta.write_text(
        "Alias;Provincia;Municipio_oficial;Codigo_INE;Fuente;Confianza\n"
        "Ciudad Real;Madrid;Madrid;28079;https://www.boe.es/;ALTA\n", encoding="utf-8"
    )
    propuesta = [{"administracion_detectada": "Ayuntamiento de Ciudad Real", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1}]
    sede = resolver_sedes(propuesta, StringIO(), ruta_alias=ruta)["sedes"][0]
    assert (sede["Municipio"], sede["Provincia"]) == ("Ciudad Real", "Ciudad Real")
    assert "CATALOGO_ALIAS" not in sede["metodo_resolucion"]


def test_diagnostico_alias_cuenta_alias_y_no_aproxima_el_pendiente():
    propuestas = resolver_sedes([
        {"administracion_detectada": "Ayuntamiento de Palma de Mallorca (Illes Balears)", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 2},
        {"administracion_detectada": "Ayuntamiento de Granada", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ], StringIO())["propuestas"]
    diagnostico = analizar_alias_municipios(propuestas, total_base=3)
    assert diagnostico["filas_resueltas_por_alias"] == 2
    assert diagnostico["ranking_municipios_no_resueltos"] == []
    assert diagnostico["umbrales_cobertura"]["50"]["aliases_necesarios"] == 1


def test_granada_y_la_granada_se_resuelven_sin_eliminar_articulos():
    propuestas = [
        {"administracion_detectada": "Ayuntamiento de Granada", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
        {"administracion_detectada": "Ayuntamiento de La Granada", "familia": "AYUNTAMIENTO", "confianza": "ALTA", "titulo": "real", "filas_oposiciones": 1},
    ]
    sedes = resolver_sedes(propuestas, StringIO())["sedes"]
    assert [(x["Municipio"], x["Provincia"]) for x in sedes] == [("Granada", "Granada"), ("Granada, La", "Barcelona")]


def test_preparar_actualizaciones_separa_administracion_de_sede(tmp_path):
    excel = tmp_path / "datos.xlsx"; _excel_prueba(excel)
    cambios = preparar_actualizaciones(excel, _resultado_prueba())
    assert [x["valores"] for x in cambios] == [
        {"Administración": "Universidad de Salamanca"},
        {"Administración": "Ayuntamiento de Ciudad Real", "Municipio": "Ciudad Real", "Provincia": "Ciudad Real"},
    ]


def test_escritura_temporal_solo_modifica_campos_autorizados_y_crea_backup(tmp_path):
    excel = tmp_path / "datos.xlsx"; _excel_prueba(excel)
    cambios = preparar_actualizaciones(excel, _resultado_prueba())
    backup, firma = _crear_backup_enriquecimiento(excel, tmp_path / "backups")
    assert backup.exists() and firma["sha256"] == __import__("enriquecer_administraciones_historicas").firma_excel(excel)["sha256"]
    _escribir_actualizaciones_atomicas(excel, cambios, StringIO())
    libro = load_workbook(excel, read_only=True, data_only=False)
    try:
        filas = list(libro["Oposiciones"].iter_rows(values_only=True))
        assert filas[1][1:4] == ("Universidad de Salamanca", None, None)
        assert filas[2][1:4] == ("Ayuntamiento de Ciudad Real", "Ciudad Real", "Ciudad Real")
        assert filas[1][5:] == ("Puesto 1", 1, "--")
        assert list(libro["Búsquedas"].iter_rows(values_only=True)) == [("Código", "Texto"), ("x", "sin cambios")]
    finally:
        libro.close()


def test_validacion_temporal_rechaza_cambio_no_autorizado(tmp_path):
    origen = tmp_path / "origen.xlsx"; temporal = tmp_path / "temporal.xlsx"; _excel_prueba(origen); _excel_prueba(temporal)
    libro = load_workbook(temporal); libro["Oposiciones"].cell(2, 6).value = "Puesto alterado"; libro.save(temporal); libro.close()
    with pytest.raises(RuntimeError, match="Cambio no autorizado"):
        _validar_temporal(origen, temporal, [])


def test_aplicar_idempotente_no_crea_backup_si_no_hay_cambios(tmp_path, monkeypatch):
    excel = tmp_path / "datos.xlsx"; _excel_prueba(excel)
    import enriquecer_administraciones_historicas as modulo
    monkeypatch.setattr(modulo, "calcular_propuestas", lambda *args, **kwargs: _resultado_prueba())
    monkeypatch.setattr(modulo, "preparar_actualizaciones", lambda *args, **kwargs: [])
    resultado = aplicar_enriquecimiento(excel, backup_directorio=tmp_path / "backups", stream=StringIO())
    assert resultado["ya_aplicado"] is True
    assert not (tmp_path / "backups").exists()


def test_fallo_backup_y_fallo_validacion_conservan_el_original(tmp_path, monkeypatch):
    excel = tmp_path / "datos.xlsx"; _excel_prueba(excel)
    import enriquecer_administraciones_historicas as modulo
    firma_original = modulo.firma_excel(excel)
    monkeypatch.setattr(modulo.shutil, "copy2", lambda *args, **kwargs: (_ for _ in ()).throw(OSError("fallo backup")))
    with pytest.raises(OSError, match="fallo backup"):
        _crear_backup_enriquecimiento(excel, tmp_path / "backups")
    assert modulo.firma_excel(excel) == firma_original
    monkeypatch.undo()
    cambios = preparar_actualizaciones(excel, _resultado_prueba())
    monkeypatch.setattr(modulo, "_validar_temporal", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("fallo validación")))
    with pytest.raises(RuntimeError, match="fallo validación"):
        _escribir_actualizaciones_atomicas(excel, cambios, StringIO())
    assert modulo.firma_excel(excel) == firma_original
