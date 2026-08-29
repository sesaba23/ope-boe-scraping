"""Diagnóstico experimental, de solo lectura, de sedes de administraciones históricas."""
import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from mapa_plazas import _variantes_nombre_catalogo, normalizar_nombre_municipal


FAMILIAS = (
    ("AYUNTAMIENTO", re.compile(r"\bayuntamiento\s+de(?:\s+la|l|l\s+)?\s+(.+?)(?=\s*[,.;]|$)", re.I)),
    ("DIPUTACION_PROVINCIAL", re.compile(r"\bdiputaci[oó]n\s+provincial\s+de\s+(.+?)(?=\s*[,.;]|$)", re.I)),
    ("DIPUTACION", re.compile(r"\bdiputaci[oó]n\s+de\s+(.+?)(?=\s*[,.;]|$)", re.I)),
    ("UNIVERSIDAD", re.compile(r"\buniversidad\s+de\s+(.+?)(?=\s*[,.;]|$)", re.I)),
)
VERSION_HISTORICA = "historico-experimental-2004"
RUTA_CATALOGO_OFICIAL = Path(__file__).parent / "datos" / "municipios_oficial.csv"


def cargar_catalogo(ruta=None):
    ruta = ruta or (RUTA_CATALOGO_OFICIAL if RUTA_CATALOGO_OFICIAL.exists()
                    else Path(__file__).parent / "assets" / "resources" / "municipios.csv")
    return pd.read_csv(ruta, sep=";")


def crear_indice_municipios(catalogo):
    """Índice exacto con las variantes seguras ya usadas por mapa_plazas."""
    indice = defaultdict(list)
    for fila in catalogo.to_dict(orient="records"):
        for variante in _variantes_nombre_catalogo(fila["Población"]):
            indice[normalizar_nombre_municipal(variante)].append(fila)
    return indice


def detectar_administracion(texto):
    """Extrae solo las cuatro familias explícitas de alta confianza."""
    if not isinstance(texto, str) or not texto.strip():
        return None
    for tipo, patron in FAMILIAS:
        coincidencia = patron.search(texto)
        if coincidencia:
            entidad = " ".join(coincidencia.group(1).split()).strip(" -")
            if entidad:
                etiqueta = {
                    "AYUNTAMIENTO": "Ayuntamiento de",
                    "DIPUTACION_PROVINCIAL": "Diputación Provincial de",
                    "DIPUTACION": "Diputación de",
                    "UNIVERSIDAD": "Universidad de",
                }[tipo]
                return {"tipo_administracion": tipo, "administracion_detectada": f"{etiqueta} {entidad}",
                        "entidad_extraida": entidad}
    return None


def _entidad_y_provincia(entidad):
    coincidencia = re.fullmatch(r"\s*(.*?)\s*\(([^()]+)\)\s*", entidad)
    if not coincidencia:
        return entidad.strip(), None
    return coincidencia.group(1).strip(), coincidencia.group(2).strip()


def _variantes_provincia(nombre):
    """Equivalencias explícitas de grafías cooficiales separadas por '/'."""
    texto = str(nombre).strip()
    return ({normalizar_nombre_municipal(texto), normalizar_nombre_municipal(texto.replace("/", " "))}
            | {normalizar_nombre_municipal(parte) for parte in texto.split("/") if parte.strip()})


def resolver_entidad(entidad, indice):
    """Resuelve solo coincidencias normalizadas exactas e inequívocas."""
    nombre, provincia_indicada = _entidad_y_provincia(entidad)
    filas = indice.get(normalizar_nombre_municipal(nombre), [])
    if provincia_indicada:
        provincias = _variantes_provincia(provincia_indicada)
        por_provincia = [fila for fila in filas if _variantes_provincia(fila["Provincia"]) & provincias]
        if por_provincia:
            filas = por_provincia
    unicas = {(fila["Población"], fila["Provincia"]) for fila in filas}
    if len(unicas) == 1:
        municipio, provincia = next(iter(unicas))
        return municipio, provincia, "CATALOGO_MUNICIPIOS_EXACTO", "ALTA"
    if len(unicas) > 1:
        return "", "", "CATALOGO_MUNICIPIOS_AMBIGUO", "AMBIGUA"
    return "", "", "CATALOGO_MUNICIPIOS_SIN_COINCIDENCIA", "NO_RESUELTA"


def _es_historica(serie):
    return serie.astype(str).eq(VERSION_HISTORICA)


def _leer_hoja_columnas(libro, nombre_hoja, columnas):
    """Lee solo las columnas diagnósticas desde un único libro en modo lectura."""
    hoja = libro[nombre_hoja]
    filas = hoja.iter_rows(values_only=True)
    cabeceras = next(filas)
    indices = {nombre: cabeceras.index(nombre) for nombre in columnas}
    return pd.DataFrame(
        ({nombre: fila[indices[nombre]] for nombre in columnas} for fila in filas),
        columns=columnas,
    )


def cargar_historicos(excel="BOE-oposiciones.xlsx"):
    libro = load_workbook(excel, read_only=True, data_only=True)
    try:
        publicaciones = _leer_hoja_columnas(
            libro, "Publicaciones", ["Publicacion_ID", "Titulo_original", "Version_extractor", "Fecha_BOE", "Enlace"])
        oposiciones = _leer_hoja_columnas(
            libro, "Oposiciones", ["Publicacion_ID", "Administración", "Version_extractor"])
    finally:
        libro.close()
    return (publicaciones[_es_historica(publicaciones["Version_extractor"])].copy(),
            oposiciones[_es_historica(oposiciones["Version_extractor"])].copy())


def diagnosticar(excel="BOE-oposiciones.xlsx", catalogo=None, publicaciones=None, oposiciones=None):
    if publicaciones is None or oposiciones is None:
        publicaciones, oposiciones = cargar_historicos(excel)
    catalogo = cargar_catalogo(catalogo)
    indice = crear_indice_municipios(catalogo)
    pubs, opos = publicaciones, oposiciones
    titulos = pubs.set_index("Publicacion_ID")["Titulo_original"].to_dict()
    filas_por_id = defaultdict(list)
    for fila in opos.to_dict(orient="records"):
        filas_por_id[fila.get("Publicacion_ID")].append(fila)
    resultados, detectados_titulo = [], set()
    for publicacion_id, titulo in titulos.items():
        dato = detectar_administracion(titulo)
        if dato:
            detectados_titulo.add(publicacion_id)
            municipio, provincia, metodo, confianza = resolver_entidad(dato["entidad_extraida"], indice)
            resultados.append({"Publicacion_ID": publicacion_id, "titulo_original": titulo,
                               **dato, "municipio_resuelto": municipio, "provincia_resuelta": provincia,
                               "metodo_fuente": f"TITULO_BOE+{metodo}", "confianza": confianza,
                               "alcance_filas": "PUBLICACION_COMPLETA"})
            continue
        # Fuente secundaria: la Administración ya almacenada en filas históricas.
        for administracion in sorted({str(x.get("Administración") or "") for x in filas_por_id.get(publicacion_id, [])}):
            dato = detectar_administracion(administracion)
            if dato:
                municipio, provincia, metodo, confianza = resolver_entidad(dato["entidad_extraida"], indice)
                resultados.append({"Publicacion_ID": publicacion_id, "titulo_original": titulo if isinstance(titulo, str) else "",
                                   **dato, "municipio_resuelto": municipio, "provincia_resuelta": provincia,
                                   "metodo_fuente": f"ADMINISTRACION_EXISTENTE+{metodo}", "confianza": confianza,
                                   "alcance_filas": "FILAS_CON_ADMINISTRACION_COINCIDENTE"})
    ids_titulo = {x["Publicacion_ID"] for x in resultados if x["metodo_fuente"].startswith("TITULO_BOE")}
    claves_admin = {(x["Publicacion_ID"], x["administracion_detectada"]) for x in resultados if x["metodo_fuente"].startswith("ADMINISTRACION_EXISTENTE")}
    filas_potenciales = 0
    for fila in opos.to_dict(orient="records"):
        if fila.get("Publicacion_ID") in ids_titulo or (fila.get("Publicacion_ID"), str(fila.get("Administración") or "")) in claves_admin:
            filas_potenciales += 1
    titulo_util = sum(isinstance(x, str) and bool(x.strip()) for x in pubs["Titulo_original"])
    resumen = {
        "publicaciones_analizadas": len(pubs), "filas_oposiciones_analizadas": len(opos),
        "titulos_disponibles": titulo_util, "titulos_sin_informacion_utilizable": len(pubs) - titulo_util,
        "detecciones_por_familia": dict(Counter(x["tipo_administracion"] for x in resultados)),
        "administraciones_unicas": len({x["administracion_detectada"] for x in resultados}),
        "municipios_resueltos": sum(bool(x["municipio_resuelto"]) for x in resultados),
        "provincias_resueltas": sum(bool(x["provincia_resuelta"]) for x in resultados),
        "no_resueltos": sum(x["confianza"] == "NO_RESUELTA" for x in resultados),
        "ambiguedades": sum(x["confianza"] == "AMBIGUA" for x in resultados),
        "filas_potencialmente_geolocalizables": filas_potenciales,
        "porcentaje_filas_potencial": round(100 * filas_potenciales / len(opos), 2) if len(opos) else 0,
    }
    frecuentes = Counter(str(x.get("Administración") or "") for x in opos.to_dict(orient="records"))
    return {"resumen": resumen, "resultados": resultados,
            "administraciones_mas_frecuentes": [{"administracion": nombre, "filas": numero}
                                                   for nombre, numero in frecuentes.most_common(20)]}


def escribir_informes(datos, directorio="informes/diagnostico_administraciones_historicas"):
    destino = Path(directorio); destino.mkdir(parents=True, exist_ok=True)
    (destino / "diagnostico_administraciones_historicas.json").write_text(
        json.dumps(datos, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    r = datos["resumen"]
    lineas = ["# Diagnóstico de administraciones históricas", "", "## Resumen", ""]
    lineas += [f"- {k.replace('_', ' ')}: {v}" for k, v in r.items()]
    lineas += ["", "## Ejemplos por familia", ""]
    for familia in sorted({x["tipo_administracion"] for x in datos["resultados"]}):
        ejemplo = next(x for x in datos["resultados"] if x["tipo_administracion"] == familia)
        lineas += [f"- {familia}: {ejemplo['Publicacion_ID']} — {ejemplo['administracion_detectada']} → {ejemplo['municipio_resuelto'] or 'sin resolver'}"]
    lineas += ["", "## 20 administraciones más frecuentes", ""]
    lineas += [f"- {x['administracion']}: {x['filas']}" for x in datos["administraciones_mas_frecuentes"]]
    (destino / "diagnostico_administraciones_historicas.md").write_text("\n".join(lineas) + "\n", encoding="utf-8")


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", default="BOE-oposiciones.xlsx")
    parser.add_argument("--salida", default="informes/diagnostico_administraciones_historicas")
    args = parser.parse_args(argv)
    datos = diagnosticar(args.excel)
    escribir_informes(datos, args.salida)
    print(json.dumps(datos["resumen"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
