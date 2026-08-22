import hashlib
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import auditoria_datos


def _crear_libro(ruta, opcionales=True):
    enlace1 = "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-1001"
    oposiciones = pd.DataFrame(
        [
            {
                "Num_plazas": 2, "Puesto": "Ingeniero/a Técnico", "Administración": "Ayuntamiento Ávila",
                "Escala": "Técnica", "Subescala": "--", "Clase": "No disponible",
                "Sistema": "Concurso-Oposición", "Turno": "Libre", "Fecha_boe": "10 de agosto de 2026",
                "Publicación": "Título", "Enlace": enlace1, "Municipio": "Ávila", "Provincia": "Ávila",
                "Latitud": 40.65, "Longitud": -4.69, "Habitantes": 58000,
                "Publicacion_ID": "BOE-A-2026-1001", "Version_extractor": "1",
                "Fecha_analisis": "2026-08-10 12:00:00",
            },
            {
                "Num_plazas": 2, "Puesto": "Ingeniero/a Técnico", "Administración": "Ayuntamiento Ávila",
                "Escala": "Técnica", "Subescala": "--", "Clase": "No disponible",
                "Sistema": "Concurso-Oposición", "Turno": "Libre", "Fecha_boe": "10 de agosto de 2026",
                "Publicación": "Título", "Enlace": enlace1, "Municipio": "Ávila", "Provincia": "Ávila",
                "Latitud": 40.65, "Longitud": -4.69, "Habitantes": 58000,
                "Publicacion_ID": "BOE-A-2026-1001", "Version_extractor": "1",
                "Fecha_analisis": "2026-08-10 12:00:00",
            },
            {
                "Num_plazas": "texto", "Puesto": "Ingeniero Técnico", "Administración": "AYUNTAMIENTO AVILA",
                "Escala": "tecnica", "Subescala": "", "Clase": "--",
                "Sistema": "Concurso oposición", "Turno": "libre", "Fecha_boe": "fecha rota",
                "Publicación": "", "Enlace": "https://ejemplo.invalid", "Municipio": "Ávila", "Provincia": "",
                "Latitud": "norte", "Longitud": 300, "Habitantes": "muchos",
                "Publicacion_ID": "invalido", "Version_extractor": "legacy", "Fecha_analisis": "ayer",
            },
            {
                "Num_plazas": 1, "Puesto": "Auxiliar", "Administración": "Entidad",
                "Sistema": "--", "Turno": "--", "Fecha_boe": "11 de agosto de 2026",
                "Enlace": "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-1004",
                "Municipio": None, "Provincia": "No disponible", "Latitud": None, "Longitud": None,
                "Habitantes": None, "Publicacion_ID": "BOE-A-2026-1004",
                "Version_extractor": "1", "Fecha_analisis": None,
            },
        ]
    )
    with pd.ExcelWriter(ruta) as writer:
        oposiciones.to_excel(writer, sheet_name="Oposiciones", index=False)
        if opcionales:
            pd.DataFrame(
                [
                    {"Publicacion_ID": "BOE-A-2026-1001", "Fecha_BOE": "10 de agosto de 2026", "Fecha_ultimo_analisis": "2026-08-10 12:00:00", "Version_extractor": "1", "Estado_analisis": "con_coincidencias", "Coincidencias": 1, "Enlace": enlace1},
                    {"Publicacion_ID": "BOE-A-2026-1001", "Fecha_BOE": "11 de agosto de 2026", "Fecha_ultimo_analisis": "mal", "Version_extractor": "legacy", "Estado_analisis": "desconocido", "Coincidencias": -1, "Enlace": enlace1},
                    {"Publicacion_ID": "BOE-A-2026-1003", "Fecha_BOE": "10 de agosto de 2026", "Version_extractor": "1", "Estado_analisis": "con_coincidencias", "Coincidencias": "x"},
                    {"Publicacion_ID": "BOE-A-2026-1004", "Fecha_BOE": "11 de agosto de 2026", "Version_extractor": "1", "Estado_analisis": "sin_coincidencias", "Coincidencias": 0},
                ]
            ).to_excel(writer, sheet_name="Publicaciones", index=False)
            pd.DataFrame(
                [
                    {"Fecha": "2026-08-10", "Estado": "consultado", "Version_extractor": "1", "Numero_publicaciones": 5},
                    {"Fecha": "2026-08-10", "Estado": "raro", "Version_extractor": "legacy", "Numero_publicaciones": -1},
                    {"Fecha": "fecha rota", "Estado": "sin_edicion", "Version_extractor": "1", "Numero_publicaciones": 2},
                ]
            ).to_excel(writer, sheet_name="Cobertura", index=False)
            pd.DataFrame({"Código": [enlace1 + "_ingeniero", enlace1 + "_ingeniero", "no-asociable"]}).to_excel(writer, sheet_name="Búsquedas", index=False)
            pd.DataFrame(
                {"Tipo de error": ["HTTP", "HTTP", "Estructura"],
                 "Enlace Web": [enlace1, enlace1, "https://sin-resolver.invalid"]}
            ).to_excel(writer, sheet_name="Log-errores", index=False)


def _hash(ruta):
    return hashlib.sha256(Path(ruta).read_bytes()).hexdigest()


def test_auditoria_cubre_metricas_calidad_y_relaciones(tmp_path):
    ruta = tmp_path / "datos.xlsx"
    _crear_libro(ruta)

    resultado = auditoria_datos.auditar_datos(ruta)

    assert resultado["resumen"]["Filas de Oposiciones"] == 4
    assert resultado["resumen"]["Filas con Publicacion_ID válido"] == 3
    calidad = {fila["Columna"]: fila for fila in resultado["calidad_columnas"]}
    assert calidad["Subescala"]["--"] == 2
    assert calidad["Clase"]["No disponible"] == 2
    assert resultado["geolocalizacion"]["Coordenadas no numéricas"] == 1
    assert resultado["geolocalizacion"]["Coordenadas fuera de rango"] == 1
    assert resultado["fechas"]["Fecha_boe inválidas"] == 1
    assert resultado["fechas"]["Fecha_analisis inválidas"] == 1
    assert resultado["publicaciones"]["Publicacion_ID duplicados"] == 2
    assert resultado["publicaciones"]["con_coincidencias sin Oposiciones"] == 1
    assert resultado["publicaciones"]["sin_coincidencias con Oposiciones"] == 1
    assert resultado["cobertura"]["Fechas duplicadas"] == 2
    assert resultado["cobertura"]["Estados desconocidos"] == 1
    assert resultado["busquedas"]["Códigos únicos"] == 2
    assert resultado["busquedas"]["Códigos no asociables"] == 1
    assert resultado["errores"]["Potencialmente resueltos"] == 2
    assert resultado["duplicados"]["Filas duplicadas exactas"] == 2


def test_detecta_variantes_y_rankings_sin_normalizacion_semantica(tmp_path):
    ruta = tmp_path / "datos.xlsx"
    _crear_libro(ruta)

    resultado = auditoria_datos.auditar_datos(ruta)

    assert resultado["categoricos"]["Sistema"]["variantes"]
    assert resultado["categoricos"]["Escala"]["variantes"]
    assert resultado["puestos"]["variantes"]
    assert resultado["administraciones"]["variantes"]
    assert resultado["puestos"]["top_registros"][0]["Valor"] == "Ingeniero/a Técnico"
    assert resultado["puestos"]["top_plazas"][0]["Plazas"] == 4.0


def test_hojas_opcionales_ausentes_no_impiden_auditoria(tmp_path):
    ruta = tmp_path / "solo-oposiciones.xlsx"
    _crear_libro(ruta, opcionales=False)

    resultado = auditoria_datos.auditar_datos(ruta)

    assert resultado["publicaciones"]["Filas"] == 0
    assert resultado["cobertura"]["Filas"] == 0
    assert resultado["busquedas"]["Total"] == 0
    assert resultado["errores"]["Total"] == 0


def test_genera_markdown_con_todas_las_secciones(tmp_path):
    ruta = tmp_path / "datos.xlsx"
    informe = tmp_path / "informe_auditoria_datos.md"
    _crear_libro(ruta)

    auditoria_datos.ejecutar_auditoria(ruta, informe)

    contenido = informe.read_text(encoding="utf-8")
    for seccion in [
        "Resumen ejecutivo", "Calidad por columnas", "Geolocalización",
        "Valores categóricos", "Puestos", "Administraciones", "Fechas",
        "Publicaciones", "Cobertura", "Búsquedas", "Log de errores",
        "Duplicados e inconsistencias", "Recomendaciones",
    ]:
        assert f"## {seccion}" in contenido
    assert "## Diagnóstico de segundo nivel" in contenido
    for subseccion in [
        "Versiones de Publicaciones",
        "Convocatorias aparentemente duplicadas",
        "Publicaciones multiconvocatoria",
        "Estado real del Log de errores",
        "Geolocalización pendiente por tipo de administración",
    ]:
        assert f"### {subseccion}" in contenido
    assert "HISTÓRICO" in contenido
    assert "VERSION_EXTRACTOR" not in contenido


def test_auditor_es_estrictamente_solo_lectura(tmp_path):
    ruta = tmp_path / "datos.xlsx"
    informe = tmp_path / "informe_auditoria_datos.md"
    _crear_libro(ruta)
    contenido_antes = ruta.read_bytes()
    tamaño_antes = ruta.stat().st_size
    hash_antes = _hash(ruta)
    mtime_antes = ruta.stat().st_mtime_ns
    libro_antes = load_workbook(ruta, read_only=True)
    hojas_antes = libro_antes.sheetnames
    libro_antes.close()

    auditoria_datos.ejecutar_auditoria(ruta, informe)

    assert ruta.read_bytes() == contenido_antes
    assert ruta.stat().st_size == tamaño_antes
    assert _hash(ruta) == hash_antes
    assert ruta.stat().st_mtime_ns == mtime_antes
    libro = load_workbook(ruta, read_only=True)
    assert libro.sheetnames == hojas_antes
    libro.close()
    assert set(tmp_path.iterdir()) == {ruta, informe}
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_cli_admite_excel_configurable(monkeypatch, tmp_path, capsys):
    ruta = tmp_path / "datos.xlsx"
    _crear_libro(ruta, opcionales=False)
    monkeypatch.chdir(tmp_path)

    auditoria_datos.main(["--excel", str(ruta)])

    assert (tmp_path / "informe_auditoria_datos.md").exists()
    assert "Auditoría completada" in capsys.readouterr().out


def _crear_libro_segundo_nivel(ruta):
    def fila(pid, puesto, enlace, **cambios):
        datos = {
            "Publicacion_ID": pid, "Puesto": puesto, "Enlace": enlace,
            "Num_plazas": 1, "Turno": "Libre", "Sistema": "Oposición",
            "Escala": "General", "Subescala": "Técnica", "Clase": "A1",
            "Administración": "Ayuntamiento de Uno", "Municipio": "Uno",
            "Provincia": "Madrid", "Latitud": 40.0, "Longitud": -3.0,
            "Fecha_boe": "10 de agosto de 2026",
        }
        datos.update(cambios)
        return datos

    enlace_a = "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-2001"
    enlace_b = "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-2002"
    enlace_c = "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-2003"
    oposiciones = pd.DataFrame([
        fila("BOE-A-2026-2001", "Técnico", enlace_a),
        fila("BOE-A-2026-2001", "Técnico", enlace_a, Turno="Discapacidad", Num_plazas=2),
        fila("BOE-A-2026-2002", "Auxiliar", enlace_b),
        fila("BOE-A-2026-2002", "Auxiliar", enlace_b, Administración="Ayuntamiento de Dos", Municipio="Dos", Provincia="Sevilla"),
        fila("BOE-A-2026-2003", "Arquitecto", enlace_c),
        fila("BOE-A-2026-2003", "Ingeniero", enlace_c, Sistema="Concurso", Num_plazas=3),
        fila("BOE-A-2026-2010", "Administrativo", "u10", Administración="Ayuntamiento Sin Coordenadas", Municipio=None, Provincia=None, Latitud=None, Longitud=None),
        fila("BOE-A-2026-2011", "Administrativo", "u11", Administración="Diputación Provincial de X", Municipio=None, Provincia=None, Latitud=None, Longitud=None),
        fila("BOE-A-2026-2012", "Administrativo", "u12", Administración="Cabildo Insular de X", Municipio=None, Provincia=None, Latitud=None, Longitud=None),
        fila("BOE-A-2026-2013", "Administrativo", "u13", Administración="Mancomunidad de X", Municipio=None, Provincia=None, Latitud=None, Longitud=None),
    ])
    publicaciones = pd.DataFrame([
        {"Publicacion_ID": "BOE-A-2026-2001", "Fecha_BOE": "10 de agosto de 2026", "Version_extractor": "1", "Estado_analisis": "con_coincidencias", "Coincidencias": 2, "Fecha_ultimo_analisis": "2026-08-12 10:00:00"},
        {"Publicacion_ID": "BOE-A-2026-2002", "Fecha_BOE": "10 de agosto de 2026", "Version_extractor": "legacy", "Estado_analisis": "con_coincidencias", "Coincidencias": 2},
        {"Publicacion_ID": "BOE-A-2026-2003", "Fecha_BOE": "10 de agosto de 2026", "Version_extractor": "corrupta", "Estado_analisis": "con_coincidencias", "Coincidencias": 2},
        {"Publicacion_ID": "BOE-A-2026-2004", "Fecha_BOE": "12 de agosto de 2026", "Version_extractor": "2", "Estado_analisis": "sin_coincidencias", "Coincidencias": 0},
    ])
    cobertura = pd.DataFrame([
        {"Fecha": "2026-08-11", "Estado": "consultado", "Version_extractor": "1", "Fecha_ultima_consulta": "2026-08-12 12:00:00", "Numero_publicaciones": 0},
        {"Fecha": "2026-08-12", "Estado": "error", "Version_extractor": None, "Fecha_ultima_consulta": "2026-08-12 12:00:00", "Numero_publicaciones": None},
    ])
    errores = pd.DataFrame([
        {"Fecha": "2026-08-11 09:00:00", "Tipo de error": "Documento", "Enlace Web": enlace_a},
        {"Fecha": "2026-08-11 09:00:00", "Tipo de error": "Documento", "Enlace Web": "https://www.boe.es/diario_boe/txt.php?id=BOE-A-2026-9999"},
        {"Fecha": "2026-08-11 09:00:00", "Tipo de error": "Índice", "Enlace Web": "https://www.boe.es/boe/dias/2026/08/11/index.php?s=2B"},
        {"Fecha": "2026-08-12 09:00:00", "Tipo de error": "Índice", "Enlace Web": "https://www.boe.es/boe/dias/2026/08/12/index.php?s=2B"},
    ])
    with pd.ExcelWriter(ruta) as writer:
        oposiciones.to_excel(writer, sheet_name="Oposiciones", index=False)
        publicaciones.to_excel(writer, sheet_name="Publicaciones", index=False)
        cobertura.to_excel(writer, sheet_name="Cobertura", index=False)
        errores.to_excel(writer, sheet_name="Log-errores", index=False)


def test_diagnostico_segundo_nivel_reclasifica_hallazgos(tmp_path):
    ruta = tmp_path / "segundo-nivel.xlsx"
    _crear_libro_segundo_nivel(ruta)

    diagnostico = auditoria_datos.auditar_datos(ruta)["diagnostico_segundo_nivel"]

    versiones = diagnostico["versiones"]
    assert versiones["Legacy pendientes de reprocesamiento"] == 1
    assert versiones["Valores realmente inválidos"] == 1
    assert versiones["Versiones numéricas válidas"] == 1
    assert versiones["Versiones numéricas futuras"] == 1
    duplicadas = diagnostico["convocatorias_duplicadas"]
    assert duplicadas["Grupos totales"] == 2
    assert duplicadas["LEGÍTIMO"] == 1
    assert duplicadas["REVISAR"] == 1
    multi = diagnostico["multiconvocatorias"]
    assert multi["NORMAL_MULTICONVOCATORIA"] == 2
    assert multi["POSIBLE_INCONSISTENCIA"] == 1
    assert multi["Falsos positivos de primera auditoría"] == 2


def test_diagnostico_errores_y_geolocalizacion_por_tipo(tmp_path):
    ruta = tmp_path / "segundo-nivel.xlsx"
    _crear_libro_segundo_nivel(ruta)

    diagnostico = auditoria_datos.auditar_datos(ruta)["diagnostico_segundo_nivel"]

    errores = diagnostico["estado_errores"]
    assert errores["RESUELTO"] == 2
    assert errores["PENDIENTE"] == 1
    assert errores["ERROR_DE_INDICE"] == 1
    geo = diagnostico["geolocalizacion_por_tipo"]
    resumen = {fila["Tipo"]: fila["Filas"] for fila in geo["resumen_por_tipo"]}
    assert resumen["Ayuntamiento"] == 1
    assert resumen["Diputación Provincial"] == 1
    assert resumen["Cabildo Insular"] == 1
    assert resumen["Mancomunidad/Mancomunitat"] == 1
    assert geo["ayuntamientos"][0]["Administración"] == "Ayuntamiento Sin Coordenadas"
