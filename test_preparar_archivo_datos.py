import warnings

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from pandas.core.strings.accessor import StringMethods

import preparar_archivo_datos
from fechas import convertir_fecha
from preparar_archivo_datos import combinar_dataframes, prepara_data_frame_mostrar_resultados


@pytest.fixture(autouse=True)
def simular_conversion_fechas(monkeypatch):
    fechas_iso = {
        "2 de enero de 2025": "2025/01/02",
        "3 de enero de 2025": "2025/01/03",
    }
    monkeypatch.setattr(
        preparar_archivo_datos, "convertir_fecha", lambda fecha: fechas_iso[fecha]
    )


def crear_dataframe():
    return pd.DataFrame(
        [
            {"Fecha_boe": "2 de enero de 2025", "Puesto": "Ingeniero Industrial"},
            {"Fecha_boe": "3 de enero de 2025", "Puesto": "Auxiliar Administrativo"},
        ]
    )


def crear_convocatoria(**cambios):
    convocatoria = {
        "Puesto": "Ingeniero",
        "Fecha_boe": "2 de enero de 2025",
        "Administración": "Ayuntamiento de Ejemplo",
        "Enlace": "https://www.boe.es/ejemplo",
        "Num_plazas": 1,
        "Turno": "Libre",
        "Sistema": "Oposición",
        "Escala": "Administración Especial",
        "Subescala": "Técnica",
        "Clase": "Superior",
    }
    convocatoria.update(cambios)
    return convocatoria


def combinar_convocatorias(*convocatorias):
    return combinar_dataframes(
        pd.DataFrame(convocatorias).to_dict(orient="list"),
        {"Código": []},
        pd.DataFrame(),
        pd.DataFrame({"Código": []}),
    )[0]


def test_combinar_dataframes_elimina_filas_identicas():
    convocatoria = crear_convocatoria()

    resultado = combinar_convocatorias(convocatoria, convocatoria.copy())

    assert len(resultado) == 1


def test_combinar_dataframes_conserva_convocatorias_con_distinto_turno():
    resultado = combinar_convocatorias(
        crear_convocatoria(), crear_convocatoria(Turno="Discapacidad")
    )

    assert len(resultado) == 2


def test_combinar_dataframes_conserva_convocatorias_con_distinto_numero_plazas():
    resultado = combinar_convocatorias(
        crear_convocatoria(), crear_convocatoria(Num_plazas=2)
    )

    assert len(resultado) == 2


@pytest.mark.parametrize(
    ("campo", "valor"),
    [
        ("Sistema", "Concurso-oposición"),
        ("Escala", "Administración General"),
        ("Subescala", "Administrativa"),
        ("Clase", "Auxiliar"),
    ],
)
def test_combinar_dataframes_conserva_diferencias_en_campos_identificativos(
    campo, valor
):
    resultado = combinar_convocatorias(
        crear_convocatoria(), crear_convocatoria(**{campo: valor})
    )

    assert len(resultado) == 2


def test_filtrado_no_modifica_dataframe_original_y_mantiene_resultado():
    df_original = crear_dataframe()

    resultado = prepara_data_frame_mostrar_resultados(
        "ingeniero industrial", df_original, ["2025/01/01", "2025/01/03"]
    )

    assert "Fecha_dt" not in df_original.columns
    assert resultado["Puesto"].tolist() == ["Ingeniero Industrial"]


def test_filtrado_con_objetos_datetime_reales(monkeypatch):
    monkeypatch.setattr(preparar_archivo_datos, "convertir_fecha", convertir_fecha)
    df_original = crear_dataframe()

    resultado = prepara_data_frame_mostrar_resultados(
        "ingeniero industrial", df_original, ["2025/01/01", "2025/01/03"]
    )

    assert pd.api.types.is_datetime64_any_dtype(resultado["Fecha_dt"])
    assert resultado["Puesto"].tolist() == ["Ingeniero Industrial"]


@pytest.mark.parametrize("fecha_invalida", ["--", "", "fecha corrupta"])
def test_filtrado_excluye_fechas_invalidas_sin_modificar_fecha_original(
    monkeypatch, fecha_invalida
):
    monkeypatch.setattr(preparar_archivo_datos, "convertir_fecha", convertir_fecha)
    df_original = pd.DataFrame(
        [
            {"Fecha_boe": "2 de enero de 2025", "Puesto": "Ingeniero válido"},
            {"Fecha_boe": fecha_invalida, "Puesto": "Ingeniero inválido"},
        ]
    )

    resultado = prepara_data_frame_mostrar_resultados(
        "ingeniero", df_original, ["2025/01/01", "2025/01/03"]
    )

    assert resultado["Puesto"].tolist() == ["Ingeniero válido"]
    assert df_original["Fecha_boe"].tolist() == ["2 de enero de 2025", fecha_invalida]
    assert "Fecha_dt" not in df_original.columns


def test_filtrado_no_modifica_configuracion_global_de_warnings():
    filtros_antes = list(warnings.filters)

    prepara_data_frame_mostrar_resultados(
        "ingeniero", crear_dataframe(), ["2025/01/01", "2025/01/03"]
    )

    assert warnings.filters == filtros_antes


def test_warnings_se_restauran_si_el_filtrado_lanza_excepcion(monkeypatch):
    filtros_antes = list(warnings.filters)

    def lanzar_error(*args, **kwargs):
        raise RuntimeError("Error simulado durante el filtrado")

    monkeypatch.setattr(StringMethods, "contains", lanzar_error)

    with pytest.raises(RuntimeError, match="Error simulado durante el filtrado"):
        prepara_data_frame_mostrar_resultados(
            "ingeniero", crear_dataframe(), ["2025/01/01", "2025/01/03"]
        )

    assert warnings.filters == filtros_antes


def _crear_excel_original(ruta):
    libro = Workbook()
    libro.active.title = "Búsquedas"
    libro.create_sheet("Oposiciones")
    libro.create_sheet("Log-errores")
    libro["Oposiciones"]["A1"] = "contenido original"
    libro.save(ruta)


def _dataframes_excel():
    return (
        pd.DataFrame([crear_convocatoria()]),
        pd.DataFrame({"Código": ["codigo-1"]}),
        pd.DataFrame(
            {
                "Fecha": ["2025-01-01 12:00:00"],
                "Tipo de error": ["Error de prueba"],
                "Enlace Web": ["https://www.boe.es/error"],
            }
        ),
    )


def test_guardar_excel_escribe_temporal_y_sustituye_al_final(monkeypatch, tmp_path):
    monkeypatch.chdir(tmp_path)
    original = tmp_path / "BOE-oposiciones.xlsx"
    _crear_excel_original(original)
    reemplazos = []
    replace_real = preparar_archivo_datos.os.replace

    def reemplazar(origen, destino):
        reemplazos.append((origen, destino, preparar_archivo_datos.Path(origen).exists()))
        replace_real(origen, destino)

    monkeypatch.setattr(preparar_archivo_datos.os, "replace", reemplazar)

    preparar_archivo_datos.guardar_excel(*_dataframes_excel())

    assert len(reemplazos) == 1
    origen, destino, temporal_existia = reemplazos[0]
    assert preparar_archivo_datos.Path(origen).parent == tmp_path
    assert preparar_archivo_datos.Path(destino).name == "BOE-oposiciones.xlsx"
    assert temporal_existia
    assert not list(tmp_path.glob(".BOE-oposiciones-*.tmp.xlsx"))


def test_guardar_excel_conserva_hojas_y_contenido_esperados(monkeypatch, tmp_path):
    monkeypatch.chdir(tmp_path)
    original = tmp_path / "BOE-oposiciones.xlsx"
    _crear_excel_original(original)

    preparar_archivo_datos.guardar_excel(*_dataframes_excel())

    libro = load_workbook(original)
    assert libro.sheetnames == ["Búsquedas", "Oposiciones", "Log-errores"]
    assert libro["Búsquedas"].sheet_state == "hidden"
    libro.close()
    hojas = pd.read_excel(original, sheet_name=None)
    assert hojas["Búsquedas"]["Código"].tolist() == ["codigo-1"]
    assert hojas["Oposiciones"]["Puesto"].tolist() == ["Ingeniero"]
    assert hojas["Log-errores"]["Tipo de error"].tolist() == ["Error de prueba"]


def test_fallo_de_escritura_conserva_original_y_elimina_temporal(
    monkeypatch, tmp_path
):
    monkeypatch.chdir(tmp_path)
    original = tmp_path / "BOE-oposiciones.xlsx"
    _crear_excel_original(original)
    contenido_original = original.read_bytes()

    monkeypatch.setattr(
        pd.DataFrame,
        "to_excel",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("fallo de escritura")),
    )

    with pytest.raises(RuntimeError, match="fallo de escritura"):
        preparar_archivo_datos.guardar_excel(*_dataframes_excel())

    assert original.read_bytes() == contenido_original
    assert not list(tmp_path.glob(".BOE-oposiciones-*.tmp.xlsx"))


def test_fallo_de_formateo_conserva_original_y_elimina_temporal(
    monkeypatch, tmp_path
):
    monkeypatch.chdir(tmp_path)
    original = tmp_path / "BOE-oposiciones.xlsx"
    _crear_excel_original(original)
    contenido_original = original.read_bytes()
    monkeypatch.setattr(
        preparar_archivo_datos,
        "formatear_hoja_oposiciones",
        lambda *args: (_ for _ in ()).throw(RuntimeError("fallo de formato")),
    )

    with pytest.raises(RuntimeError, match="fallo de formato"):
        preparar_archivo_datos.guardar_excel(*_dataframes_excel())

    assert original.read_bytes() == contenido_original
    assert not list(tmp_path.glob(".BOE-oposiciones-*.tmp.xlsx"))


def test_segunda_ejecucion_no_adquiere_el_mismo_bloqueo(tmp_path):
    archivo = tmp_path / "BOE-oposiciones.xlsx"

    with preparar_archivo_datos.bloqueo_excel(archivo):
        with pytest.raises(
            preparar_archivo_datos.ExcelBloqueadoError,
            match="Ya hay otra ejecución",
        ):
            with preparar_archivo_datos.bloqueo_excel(archivo):
                pass


def test_bloqueo_se_libera_despues_de_ejecucion_normal(tmp_path):
    archivo = tmp_path / "BOE-oposiciones.xlsx"

    with preparar_archivo_datos.bloqueo_excel(archivo):
        pass

    with preparar_archivo_datos.bloqueo_excel(archivo):
        pass


def test_bloqueo_se_libera_despues_de_excepcion(tmp_path):
    archivo = tmp_path / "BOE-oposiciones.xlsx"

    with pytest.raises(RuntimeError, match="fallo durante la ejecución"):
        with preparar_archivo_datos.bloqueo_excel(archivo):
            raise RuntimeError("fallo durante la ejecución")

    with preparar_archivo_datos.bloqueo_excel(archivo):
        pass
