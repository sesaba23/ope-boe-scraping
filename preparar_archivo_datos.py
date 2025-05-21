from fechas import convertir_fecha

import pandas as pd
from openpyxl import Workbook
import os, sys
import re
from colorama import Fore
import warnings

# Para formatear el Excel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def preparar_excel_y_dataframes():
    # Cargamos un Excel con las fechas de los días que nos interesan
    # Verificar si el archivo Excel existe. Si no, crearlo
    if not os.path.exists("BOE-oposiciones.xlsx"):
        # Crear un nuevo libro de Excel
        wb = Workbook()
        # Renombrar la hoja activa a "busquedas"
        ws1 = wb.active
        ws1.title = "Búsquedas"
        # ws1["A1"] = "Código"
        # Crear la segunda hoja
        ws2 = wb.create_sheet(title="Oposiciones")

        ws3 = wb.create_sheet(title="Log-errores")
        # Guardar el archivo
        wb.save("BOE-oposiciones.xlsx")

    # Cargamos el archivo Excel
    excel_file = pd.ExcelFile("BOE-oposiciones.xlsx")

    # Leer todas las hojas como diccionario de DataFrames
    dataframes_dict = {
        sheet_name: excel_file.parse(sheet_name)
        for sheet_name in excel_file.sheet_names
    }

    return dataframes_dict


def combinar_dataframes(
    diccionario_puestos, diccionario_busquedas, df_opo_guardadas, df_busquedas
):
    """
    Código para crear el DataFrame
    """
    # Convertimos el "diccionario_puestos" en un DataFrame de pandas
    df_diccionario_puestos = pd.DataFrame(diccionario_puestos)
    # Combinar el DataFrame existente (df_opo_guardadas) con el nuevo (df_diccionario_puestos)
    df_combinado = pd.concat(
        [df_opo_guardadas, df_diccionario_puestos], ignore_index=True
    )

    # Eliminar duplicados si es necesario, basándonos en columnas clave
    #   (por ejemplo, "Puesto" y "Fecha")
    df_combinado = df_combinado.drop_duplicates(
        subset=["Puesto", "Fecha_boe", "Administración", "Enlace"], keep="last"
    )

    # Ahora convierto el "diccionario_busquedas" en otro DataFrame
    df_diccionario_busquedas = pd.DataFrame(diccionario_busquedas)
    # Combinar el DataFrame existente (df_busquedas) con el nuevo (df_diccionario_puestos)
    df_busquedas_combinado = pd.concat(
        [df_diccionario_busquedas, df_busquedas], ignore_index=True
    )
    # Eliminar duplicados si es necesario, basándonos en columnas clave (por ejemplo, "Código")
    df_busquedas_combinado = df_busquedas_combinado.drop_duplicates(
        subset=["Código"], keep="last"
    )

    return df_combinado, df_busquedas_combinado


def guardar_excel(df_combinado, df_busquedas_combinado, df_log_errores):
    """
    Código para guardar los resultados en un archivo Excel
    """
    # Guardar los DataFrame en el archivo Excel creado al principio si está cerrado
    try:
        with pd.ExcelWriter(
            "BOE-oposiciones.xlsx",
            mode="a",
            engine="openpyxl",
            if_sheet_exists="replace",
        ) as writer:
            df_busquedas_combinado.to_excel(writer, sheet_name="Búsquedas", index=False)
            df_combinado.to_excel(writer, sheet_name="Oposiciones", index=False)
            df_log_errores.to_excel(writer, sheet_name="Log-errores", index=False)
        # Da formato al Excel
        formatear_hoja_oposiciones()
    except PermissionError:
        print(
            f"\n{Fore.RED}El archivo 'BOE-oposiciones.xlsx' está abierto. "
            f"Cierre el archivo y vuelva a intentarlo.{Fore.RESET}"
        )
        sys.exit(0)


def prepara_data_frame_mostrar_resultados(texto_busqueda, df_combinado, lista_fechas):
    """
    Buscamos los resultados buscados por el usuario dentro del propio
    Dataframe, incluidas las fechas de inicio y fin
    """

    # Convertir la columna "Fecha" del DataFrame al formato datetime en una nueva columna
    #   "Fecha_dt" para facilitar la comparación de fechas
    #   Si cambiamos directamente el formato de la columna fecha, al ejecutar, da un warning
    df_combinado["Fecha_dt"] = df_combinado["Fecha_boe"].apply(convertir_fecha)

    # Filtrar el DataFrame por el rango de fechas
    # Las fechas de inicio y final son el primer y último elemento de "lista_fechas"
    df_combinado_filtrado_por_fecha = df_combinado[
        (df_combinado["Fecha_dt"] >= lista_fechas[0])
        & (df_combinado["Fecha_dt"] <= lista_fechas[-1])
    ]

    # Ordenar el DataFrame por la columna "Fecha_dt" en orden cronológico inverso
    df_combinado_filtrado_por_fecha = df_combinado_filtrado_por_fecha.sort_values(
        by="Fecha_dt", ascending=True
    )

    # Eliminar la columna auxiliar "Fecha_dt" si no es necesaria
    df_combinado_filtrado = df_combinado_filtrado_por_fecha.drop(columns=["Fecha_dt"])

    palabras_busqueda = texto_busqueda.split()
    patron_regex = (
        r"\b"
        + r"\s+".join([rf"{clave}[\w/@\\]*(es)?" for clave in palabras_busqueda])
        + r"(.*?)"
    )
    # Desactivar warnings relacionados con expresiones regulares
    warnings.filterwarnings("ignore", message="This pattern is interpreted")

    # Filtrar el DataFrame por coincidencias en la columna "Puesto"
    df_filtrado_por_patron = df_combinado_filtrado_por_fecha[
        df_combinado_filtrado_por_fecha["Puesto"].str.contains(
            patron_regex, flags=re.IGNORECASE, na=False
        )
    ]
    # Restaurar los warnings después de la operación
    warnings.filterwarnings("default", message="This pattern is interpreted")

    return df_filtrado_por_patron


def formatear_hoja_oposiciones(nombre_archivo="BOE-oposiciones.xlsx"):
    wb = load_workbook(nombre_archivo)
    ws = wb["Oposiciones"]

    # Poner los títulos en negrita
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Insertar autofiltro en la fila de títulos
    ws.auto_filter.ref = ws.dimensions

    # Inmovilizar la primera fila
    ws.freeze_panes = "A2"

    # Autoajustar el ancho de las columnas y limitar "Puesto" y "Administración"
    max_width = 50
    col_enlace = None
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        header = str(col[0].value).strip().lower()
        if header == "enlace":
            col_enlace = col[0].column  # Guardar el índice de la columna Enlace
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        # Limitar ancho para "Puesto" y "Administración"
        if header in ["puesto", "administración"]:
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)
        else:
            ws.column_dimensions[col_letter].width = max_length + 2

    # Formatear columna "Habitantes" como número sin decimales y con separador de miles
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip().lower() == "habitantes":
            col_letter = get_column_letter(idx)
            for row in ws.iter_rows(
                min_row=2, min_col=idx, max_col=idx, max_row=ws.max_row
            ):
                for c in row:
                    c.number_format = "#,##0"
            break
    # Formatear la columna Enlace como hipervínculo
    if col_enlace:
        col_letter = get_column_letter(col_enlace)
        for row in ws.iter_rows(
            min_row=2, min_col=col_enlace, max_col=col_enlace, max_row=ws.max_row
        ):
            for cell in row:
                url = str(cell.value)
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

    # Colorear filas alternativamente
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if idx % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill

    # Ocultar la hoja "Búsquedas" si existe
    if "Búsquedas" in wb.sheetnames:
        wb["Búsquedas"].sheet_state = "hidden"

    # Hacer "Oposiciones" la hoja activa
    wb.active = wb.sheetnames.index("Oposiciones")

    wb.save(nombre_archivo)
